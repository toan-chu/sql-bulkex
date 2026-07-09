import yaml
from openpyxl import load_workbook

import runner


EXPORT_COLUMNS = [f"export_col_{i:02d}" for i in range(1, 33)]
IMPORT_COLUMNS = [f"import_col_{i:02d}" for i in range(1, 45)]


def column_cfg():
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": EXPORT_COLUMNS,
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": IMPORT_COLUMNS,
            },
        },
        "operator_defaults": {},
    }


def write_column_yaml(path):
    path.write_text(yaml.safe_dump(column_cfg(), sort_keys=False, allow_unicode=True), encoding="utf-8")


def build_template(tmp_path):
    output = tmp_path / "request_template.xlsx"
    runner.make_request_template_v5(column_cfg(), output)
    return load_workbook(output)


def validation_for_cell(ws, cell_ref):
    for dv in ws.data_validations.dataValidation:
        if cell_ref in dv.cells:
            return dv
    return None


def test_t23_make_template_has_four_v5_sheets(tmp_path):
    wb = build_template(tmp_path)

    assert wb.sheetnames == ["Request", "Cột Export", "Cột Import", "Tham chiếu"]


def test_t24_export_sheet_row_count_matches_export_columns(tmp_path):
    wb = build_template(tmp_path)

    assert wb["Cột Export"].max_row == len(EXPORT_COLUMNS) + 1
    assert wb["Cột Export"]["A2"].value == "export_col_01"
    assert wb["Cột Export"][f"A{len(EXPORT_COLUMNS) + 1}"].value == "export_col_32"


def test_t25_operator_headers_use_registry_display_order(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]

    assert [ws.cell(row=1, column=col).value for col in range(2, 8)] == [
        "Bằng",
        "Trong danh sách",
        "Trong khoảng",
        "Bắt đầu bằng",
        "Chứa",
        "Kết thúc bằng",
    ]


def test_t26_column_a_is_visual_only_not_locked(tmp_path):
    wb = build_template(tmp_path)

    for sheet_name in ("Cột Export", "Cột Import"):
        ws = wb[sheet_name]
        assert ws["A2"].fill.fgColor.rgb == "00F2F2F2"
        assert ws["A2"].font.italic
        assert ws["A2"].protection.locked is False


def test_t27_column_sheet_has_operator_conditional_formatting(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]

    cf_ranges = [str(item) for item in ws.conditional_formatting]
    rules = []
    for item in ws.conditional_formatting:
        rules.extend(ws.conditional_formatting[item])

    assert any("A2:I33" in item for item in cf_ranges)
    assert any(rule.type == "expression" and "COUNTA($B2:$G2)>0" in rule.formula for rule in rules)


def test_make_template_cli_reads_column_yaml(monkeypatch, tmp_path):
    column_path = tmp_path / "column.yaml"
    output_path = tmp_path / "request_template.xlsx"
    log_path = tmp_path / "runner.log"
    write_column_yaml(column_path)
    monkeypatch.setattr(runner, "COLUMN_FILE", column_path)
    monkeypatch.setattr(runner, "REQUEST_TEMPLATE_FILE", output_path)
    monkeypatch.setattr(runner, "LOG_DIR", tmp_path)
    monkeypatch.setattr(runner, "RUNNER_LOG_FILE", log_path)

    assert runner.main(["--make-template"]) == 0

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Request", "Cột Export", "Cột Import", "Tham chiếu"]
