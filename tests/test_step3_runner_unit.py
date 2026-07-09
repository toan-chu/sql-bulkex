from pathlib import Path

import psycopg2
import pytest

import runner


def write_legacy_workbook(path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    rows = [
        ("Người yêu cầu", "Ana"),
        ("Loại request", "MST"),
        ("Năm", "2025"),
        ("Tháng", "01"),
        ("Giá trị 1", "010"),
        ("Ghi chú / tên request", "legacy"),
    ]
    for index, (label, value) in enumerate(rows, 1):
        ws.cell(row=index, column=1, value=label)
        ws.cell(row=index, column=2, value=value)
    wb.save(path)
    return path


class TableCursor:
    def __init__(self, existing):
        self.existing = set(existing)
        self.rows = []

    def execute(self, query, params=None):
        if "LIKE" in str(query):
            pattern = params[1].replace("%", "")
            self.rows = [(t,) for t in sorted(self.existing) if t.startswith(pattern)]
        else:
            table = params[1]
            self.rows = [(1,)] if table in self.existing else []

    def fetchone(self):
        return self.rows[0] if self.rows else None

    def fetchall(self):
        return self.rows


class FakeConn:
    closed = False

    def __init__(self):
        self.closed_count = 0

    def cursor(self):
        class Cur:
            def close(self):
                pass

        return Cur()

    def rollback(self):
        pass

    def close(self):
        self.closed_count += 1


def parsed_request(large_confirm=""):
    return {
        "request": {
            "user": "Ana",
            "request_name": "daily",
            "year": "2025",
            "month": "01",
            "large_confirm": large_confirm,
        },
        "dataset": {
            "database": "db/name",
            "schema": "public",
            "tables": "x_y{year}_{month}",
            "columns": ["mst", "hs"],
        },
        "dataset_name": "export",
        "filters": [],
        "select_cols": ["mst", "hs"],
        "warnings": [],
    }


def patch_v5_process_boundaries(monkeypatch, tmp_path, rows, export_kind="xlsx"):
    captured = {"notes": []}

    monkeypatch.setattr(runner.portal, "get_conn", lambda conns, cfg, dbname: FakeConn())
    monkeypatch.setattr(runner, "parse_request_v6", lambda path, column_cfg, op_builder=None: parsed_request())
    monkeypatch.setattr(
        runner,
        "build_jobs_from_v6_dataset",
        lambda request, dataset_name, dataset_result, cur, op_builder=None: (
            "db/name",
            ([("job_suffix", "db/name", "SQL", [])], [("Dataset", "export")], ["mst", "hs"]),
        ),
    )
    monkeypatch.setattr(runner.portal, "count_rows", lambda cur, query, params: rows)
    monkeypatch.setattr(runner, "fetch_headers", lambda conn, query, params: ["mst", "hs"])

    def fake_xlsx(conn, query, params, headers, filepath, notes):
        captured["notes"].append(notes)
        Path(filepath).write_text("xlsx", encoding="utf-8")
        return rows

    def fake_csv(conn, query, params, headers, filepath, notes):
        captured["notes"].append(notes)
        Path(filepath).write_text("csv", encoding="utf-8")
        runner.write_note_txt(Path(filepath).with_suffix(".txt"), notes)
        return rows

    monkeypatch.setattr(runner, "export_xlsx_v5_with_note", fake_xlsx)
    monkeypatch.setattr(runner, "export_csv_with_note", fake_csv)
    return captured


def test_parse_months_all_list_and_range():
    assert runner.parse_months("all") == [f"{i:02d}" for i in range(1, 13)]
    assert runner.parse_months("1,3,5") == ["01", "03", "05"]
    assert runner.parse_months("2-4") == ["02", "03", "04"]


def test_parse_split_config_validation():
    assert runner.parse_split_config("mst:2") == ("mst", 2)
    assert runner.parse_split_config("", {"column": "mst", "chars": 2}) == ("mst", 2)
    with pytest.raises(runner.RequestError, match="N phải là số dương"):
        runner.parse_split_config("mst:x")


def test_expand_tables_records_missing_tables():
    cur = TableCursor({"x_y2025_01", "x_y2025_03"})

    existing, missing = runner.expand_tables(
        cur,
        "public",
        "x_y{year}_{month}",
        {"year": "2025", "month": "1,3,5"},
    )

    assert existing == ["x_y2025_01", "x_y2025_03"]
    assert missing == ["x_y2025_05"]


def test_legacy_v4_workbook_is_rejected_with_v5_message(tmp_path):
    request_path = write_legacy_workbook(tmp_path / "legacy.xlsx")

    with pytest.raises(runner.RequestError, match="mẫu v5"):
        runner.process_request_file(
            request_path,
            {},
            {},
            {"output_dir": str(tmp_path / "out")},
        )


def test_process_request_v5_rejects_auto_threshold_without_yes(monkeypatch, tmp_path):
    request_path = tmp_path / "request.xlsx"
    request_path.write_text("placeholder", encoding="utf-8")
    captured = patch_v5_process_boundaries(monkeypatch, tmp_path, rows=301)
    settings = {
        "output_dir": str(tmp_path / "out"),
        "max_rows_auto": 300,
        "max_rows_hard": 1000,
        "filename_pattern": "{ts}_{user}_{request}",
    }

    with pytest.raises(runner.RequestError, match="Xác nhận lớn"):
        runner.process_request_file_v5(request_path, {}, {}, settings, {})

    assert captured["notes"] == []


def test_process_request_v5_yes_runs_and_moves_processed(monkeypatch, tmp_path):
    request_path = tmp_path / "request.xlsx"
    request_path.write_text("placeholder", encoding="utf-8")
    patch_v5_process_boundaries(monkeypatch, tmp_path, rows=301)
    monkeypatch.setattr(runner, "parse_request_v6", lambda path, column_cfg, op_builder=None: parsed_request("YES"))
    settings = {
        "output_dir": str(tmp_path / "out"),
        "max_rows_auto": 300,
        "max_rows_hard": 1000,
        "filename_pattern": "{ts}_{user}_{request}",
    }

    moved = runner.process_request_file_v5(request_path, {}, {}, settings, {})

    assert moved.parent.name == "processed"
    assert list((tmp_path / "out").glob("*.xlsx"))


def test_process_request_v5_rejects_hard_cap(monkeypatch, tmp_path):
    request_path = tmp_path / "request.xlsx"
    request_path.write_text("placeholder", encoding="utf-8")
    patch_v5_process_boundaries(monkeypatch, tmp_path, rows=1001)
    monkeypatch.setattr(runner, "parse_request_v6", lambda path, column_cfg, op_builder=None: parsed_request("YES"))
    settings = {
        "output_dir": str(tmp_path / "out"),
        "max_rows_auto": 300,
        "max_rows_hard": 1000,
        "filename_pattern": "{ts}_{user}_{request}",
    }

    with pytest.raises(runner.RequestError, match="vượt ngưỡng cứng"):
        runner.process_request_file_v5(request_path, {}, {}, settings, {})


def test_process_request_v5_zero_rows_adds_note_hint(monkeypatch, tmp_path):
    request_path = tmp_path / "request.xlsx"
    request_path.write_text("placeholder", encoding="utf-8")
    captured = patch_v5_process_boundaries(monkeypatch, tmp_path, rows=0)
    settings = {
        "output_dir": str(tmp_path / "out"),
        "max_rows_auto": 300,
        "max_rows_hard": 1000,
        "filename_pattern": "{ts}_{user}_{request}",
    }

    runner.process_request_file_v5(request_path, {}, {}, settings, {})

    flattened = [item for notes in captured["notes"] for item in notes]
    assert ("Gợi ý", "0 dòng: kiểm tra giá trị lọc, đặc biệt MST 10/13 số hoặc dùng bắt đầu bằng.") in flattened


def test_process_request_v5_over_excel_limit_switches_to_csv(monkeypatch, tmp_path):
    request_path = tmp_path / "request.xlsx"
    request_path.write_text("placeholder", encoding="utf-8")
    patch_v5_process_boundaries(monkeypatch, tmp_path, rows=runner.XLSX_ROW_LIMIT + 1)
    monkeypatch.setattr(runner, "parse_request_v6", lambda path, column_cfg, op_builder=None: parsed_request("YES"))
    settings = {
        "output_dir": str(tmp_path / "out"),
        "max_rows_auto": 300,
        "max_rows_hard": runner.XLSX_ROW_LIMIT + 10,
        "filename_pattern": "{ts}_{user}_{request}",
    }

    runner.process_request_file_v5(request_path, {}, {}, settings, {})

    assert list((tmp_path / "out").glob("*.csv"))
    assert list((tmp_path / "out").glob("*.txt"))


def test_public_job_suffix_does_not_expose_db_schema():
    assert runner.public_job_suffix("db_name_public_gop2bang_KG", "db/name", "public") == "KG"
    assert runner.public_job_suffix("db_name_public_x_y2025_01", "db/name", "public") == "x_y2025_01"


def test_runner_headless_missing_password_fails_fast_without_getpass(monkeypatch, tmp_path):
    log_path = tmp_path / "runner.log"
    monkeypatch.setattr(runner, "LOG_DIR", tmp_path)
    monkeypatch.setattr(runner, "RUNNER_LOG_FILE", log_path)
    monkeypatch.setattr(runner, "load_settings", lambda: {"input_dir": str(tmp_path / "requests")})
    monkeypatch.setattr(runner.portal, "load_config", lambda: {"user": "postgres"})
    monkeypatch.setattr(runner.portal, "PASSWORD_FILE", str(tmp_path / ".password"))

    def fail_getpass(prompt):
        raise AssertionError("runner must not call getpass in headless mode")

    monkeypatch.setattr(runner.portal.getpass, "getpass", fail_getpass)

    assert runner.main(["--once"]) == 1
    assert "connection.yaml hoặc file .password" in log_path.read_text(encoding="utf-8")


def test_runner_headless_reads_password_file(monkeypatch, tmp_path):
    password_file = tmp_path / ".password"
    password_file.write_text("file-secret\n", encoding="utf-8")

    monkeypatch.setattr(runner.portal, "load_config", lambda: {"user": "postgres", "password": ""})
    monkeypatch.setattr(runner.portal, "PASSWORD_FILE", str(password_file))

    cfg = runner.load_connection_config()

    assert cfg["_password"] == "file-secret"


def test_run_once_skips_temp_and_unstable_files(monkeypatch, tmp_path):
    requests = tmp_path / "requests"
    requests.mkdir()
    (requests / "~$lock.xlsx").write_text("lock", encoding="utf-8")
    stable = write_legacy_workbook(requests / "stable.xlsx")
    unstable = write_legacy_workbook(requests / "unstable.xlsx")
    seen = []

    monkeypatch.setattr(runner, "is_file_stable", lambda path, stable_wait: Path(path) != unstable)
    monkeypatch.setattr(runner, "process_request_file", lambda path, cfg, conns, settings: seen.append(Path(path).name))

    runner.run_once(
        settings={"input_dir": str(requests), "output_dir": str(tmp_path / "out")},
        cfg={},
        stable_wait=0,
    )

    assert seen == [stable.name]


def test_run_once_db_down_keeps_request_file(monkeypatch, tmp_path):
    requests = tmp_path / "requests"
    requests.mkdir()
    request = write_legacy_workbook(requests / "request.xlsx")

    def db_down(path, cfg, conns, settings):
        raise psycopg2.OperationalError("db down")

    monkeypatch.setattr(runner, "process_request_file", db_down)

    runner.run_once(
        settings={"input_dir": str(requests), "output_dir": str(tmp_path / "out")},
        cfg={},
        stable_wait=0,
    )

    assert request.exists()


def test_run_once_renames_validation_error_in_place(monkeypatch, tmp_path):
    requests = tmp_path / "requests"
    requests.mkdir()
    request = write_legacy_workbook(requests / "request.xlsx")

    def invalid(path, cfg, conns, settings):
        raise runner.RequestError("lỗi dễ hiểu")

    monkeypatch.setattr(runner, "process_request_file", invalid)

    runner.run_once(
        settings={"input_dir": str(requests), "output_dir": str(tmp_path / "out")},
        cfg={},
        stable_wait=0,
    )

    assert not request.exists()
    rejected = requests / "[LOI]_request.xlsx"
    assert rejected.exists()
    txt = requests / "[LOI]_request.txt"
    assert txt.exists()
    assert "lỗi dễ hiểu" in txt.read_text(encoding="utf-8")


def test_run_once_twice_does_not_process_processed_file(monkeypatch, tmp_path):
    requests = tmp_path / "requests"
    requests.mkdir()
    request = write_legacy_workbook(requests / "request.xlsx")
    seen = []

    def process(path, cfg, conns, settings):
        seen.append(Path(path).name)
        return runner.move_request(path, "processed")

    monkeypatch.setattr(runner, "process_request_file", process)
    settings = {"input_dir": str(requests), "output_dir": str(tmp_path / "out")}

    runner.run_once(settings=settings, cfg={}, stable_wait=0)
    runner.run_once(settings=settings, cfg={}, stable_wait=0)

    assert seen == [request.name]
