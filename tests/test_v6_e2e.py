import csv
import sys
from pathlib import Path
from unittest.mock import patch

import pytest
import yaml
from openpyxl import load_workbook

try:
    import pgserver
except ImportError:
    pgserver = None

try:
    import psycopg2
except ImportError:
    psycopg2 = None

import runner


pytestmark = pytest.mark.skipif(
    sys.platform.startswith("win") or pgserver is None or psycopg2 is None,
    reason="pgserver E2E requires Linux-compatible pgserver and psycopg2",
)


EXPORT_COLUMNS = [
    "ma_nuoc",
    "ma_so_hang_hoa",
    "phuong_thuc_van_chuyen",
    "mst_nguoi_xuat_khau",
] + [f"export_col_{i:02d}" for i in range(1, 29)]

IMPORT_COLUMNS = [
    "ma_nuoc",
    "ma_so_hang_hoa",
    "phuong_thuc_van_chuyen",
    "mst_nguoi_nhap_khau",
] + [f"import_col_{i:02d}" for i in range(1, 41)]


@pytest.fixture
def pg_uri(tmp_path):
    if not hasattr(pgserver, "get_server"):
        pytest.skip("pgserver.get_server API is unavailable")
    pg = pgserver.get_server(str(tmp_path / "pgdata"), cleanup_mode="stop")
    try:
        yield pg.get_uri()
    finally:
        pg.cleanup()


@pytest.fixture
def workspace_settings(tmp_path):
    root = tmp_path / "workspace"
    pending = root / "01_Pending"
    approved = root / "02_Approved"
    output = root / "03_Output"
    log_dir = tmp_path / "log"
    for path in (pending, approved, output, log_dir):
        path.mkdir(parents=True)
    return {
        "folders": {
            "pending": str(pending),
            "approved": str(approved),
            "output": str(output),
        },
        "log": {
            "requests_csv": str(log_dir / "requests.csv"),
            "runner_log": str(log_dir / "runner.log"),
            "portal_log": str(log_dir / "portal.log"),
        },
        "filename_pattern": "{ts}_{user}_{request}",
        "max_rows_auto": 300000,
        "max_rows_hard": 3000000,
    }


def column_cfg(columns_ready=True):
    export_columns = EXPORT_COLUMNS if columns_ready else []
    import_columns = IMPORT_COLUMNS if columns_ready else []
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": export_columns,
                "cardinality_cache": {},
                "value_cache": {},
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": import_columns,
                "cardinality_cache": {},
                "value_cache": {},
            },
        },
        "operator_defaults": {},
        "cardinality": {
            "threshold": 30,
            "sample_size": 1000,
            "skip_text_length": 100,
            "skip_columns": [],
        },
    }


def create_table(cur, schema, table, columns):
    col_sql = ", ".join(f"{column} text" for column in columns)
    cur.execute(f'CREATE TABLE "{schema}"."{table}" ({col_sql})')


def insert_rows(cur, schema, table, columns, rows):
    placeholders = ", ".join(["%s"] * len(columns))
    quoted_columns = ", ".join(f'"{column}"' for column in columns)
    query = f'INSERT INTO "{schema}"."{table}" ({quoted_columns}) VALUES ({placeholders})'
    cur.executemany(query, rows)


def make_sample_row(columns, index, dataset):
    countries = ["CN", "KR", "JP", "VN", "US"]
    hs_codes = ["8471", "847130", "8471300010", "8471300020", "8436000010"]
    modes = ["Air", "Ocean", "Rail"]
    row = {}
    for column in columns:
        row[column] = f"{dataset}_{column}_{index:02d}"
    row["ma_nuoc"] = countries[index % len(countries)]
    row["ma_so_hang_hoa"] = hs_codes[index % len(hs_codes)]
    row["phuong_thuc_van_chuyen"] = modes[index % len(modes)]
    if "mst_nguoi_xuat_khau" in row:
        row["mst_nguoi_xuat_khau"] = "0101234567" if index % 2 else "0101234567890"
    if "mst_nguoi_nhap_khau" in row:
        row["mst_nguoi_nhap_khau"] = "0301234567" if index % 2 else "0301234567890"
    if index in (7, 13):
        row["phuong_thuc_van_chuyen"] = None
    return [row[column] for column in columns]


@pytest.fixture
def seeded_pg(pg_uri):
    conn = psycopg2.connect(pg_uri)
    conn.autocommit = True
    with conn.cursor() as cur:
        cur.execute("CREATE SCHEMA vietnam_export")
        cur.execute("CREATE SCHEMA vietnam_import")
        create_table(cur, "vietnam_export", "x_y2026_06", EXPORT_COLUMNS)
        create_table(cur, "vietnam_import", "i_y2026_06", IMPORT_COLUMNS)
        insert_rows(
            cur,
            "vietnam_export",
            "x_y2026_06",
            EXPORT_COLUMNS,
            [make_sample_row(EXPORT_COLUMNS, index, "export") for index in range(50)],
        )
        insert_rows(
            cur,
            "vietnam_import",
            "i_y2026_06",
            IMPORT_COLUMNS,
            [make_sample_row(IMPORT_COLUMNS, index, "import") for index in range(50)],
        )
    conn.close()
    return pg_uri


def patch_pg_connect(monkeypatch, pg_uri):
    monkeypatch.setattr(runner.portal, "connect", lambda cfg, dbname: psycopg2.connect(pg_uri))


def by_column_row(ws):
    return {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}


def write_request(path, cfg, bang="export", request_name="e2e", requester="Hoa"):
    runner.make_request_template_v5(cfg, path)
    wb = load_workbook(path)
    req = wb["Request"]
    req["B1"] = requester
    req["B2"] = bang
    req["B3"] = "2026"
    req["B4"] = "06"
    req["B7"] = request_name
    wb.properties.lastModifiedBy = "VSTREAM\\hoa.nguyen"
    return wb


def finish_workbook(wb, path):
    wb.save(path)
    wb.close()
    return path


def read_log_rows(settings):
    with open(settings["log"]["requests_csv"], newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def defined_name_text(wb, name):
    defined = wb.defined_names.get(name)
    if isinstance(defined, list):
        defined = defined[0]
    return defined.attr_text if defined is not None else None


def test_t65_full_export_flow_success(monkeypatch, tmp_path, workspace_settings, seeded_pg):
    cfg = column_cfg()
    patch_pg_connect(monkeypatch, seeded_pg)
    monkeypatch.setattr(runner, "load_column_config", lambda: cfg)
    request_path = Path(workspace_settings["folders"]["approved"]) / "request_export.xlsx"
    wb = write_request(request_path, cfg, bang="export", request_name="export-flow")
    ws = wb["Cột Export"]
    rows = by_column_row(ws)
    ws.cell(row=rows["ma_nuoc"], column=3, value="CN, KR")
    ws.cell(row=rows["ma_so_hang_hoa"], column=5, value="8471")
    ws.cell(row=rows["ma_so_hang_hoa"], column=8, value="4")
    ws.cell(row=rows["phuong_thuc_van_chuyen"], column=9, value="YES")
    finish_workbook(wb, request_path)

    processed = runner.run_once(settings=workspace_settings, cfg={"_password": ""}, stable_wait=0)

    assert processed == 1
    outputs = list(Path(workspace_settings["folders"]["output"]).glob("*.xlsx"))
    assert len(outputs) == 1
    wb_out = load_workbook(outputs[0], read_only=True)
    try:
        assert wb_out.sheetnames == ["Data", "NOTE"]
        rows_out = list(wb_out["Data"].iter_rows(values_only=True))
        assert len(rows_out) > 1
        assert rows_out[0] == ("ma_nuoc", "ma_so_hang_hoa", "phuong_thuc_van_chuyen")
    finally:
        wb_out.close()
    assert (Path(workspace_settings["folders"]["approved"]) / "[DONE] request_export.xlsx").exists()
    [log_row] = read_log_rows(workspace_settings)
    assert log_row["status"] == "success"
    assert log_row["dataset"] == "export"
    assert int(log_row["row_count"]) > 0


def test_t66_both_flow_writes_export_import_note(monkeypatch, tmp_path, workspace_settings, seeded_pg):
    cfg = column_cfg()
    patch_pg_connect(monkeypatch, seeded_pg)
    monkeypatch.setattr(runner, "load_column_config", lambda: cfg)
    request_path = Path(workspace_settings["folders"]["approved"]) / "request_both.xlsx"
    wb = write_request(request_path, cfg, bang="both", request_name="both-flow")
    export_ws = wb["Cột Export"]
    export_rows = by_column_row(export_ws)
    export_ws.cell(row=export_rows["ma_nuoc"], column=3, value="CN, KR")
    export_ws.cell(row=export_rows["ma_so_hang_hoa"], column=5, value="8471")
    export_ws.cell(row=export_rows["ma_so_hang_hoa"], column=8, value="4")
    export_ws.cell(row=export_rows["phuong_thuc_van_chuyen"], column=9, value="YES")
    import_ws = wb["Cột Import"]
    import_rows = by_column_row(import_ws)
    import_ws.cell(row=import_rows["ma_nuoc"], column=2, value="CN")
    import_ws.cell(row=import_rows["phuong_thuc_van_chuyen"], column=9, value="YES")
    finish_workbook(wb, request_path)

    processed = runner.run_once(settings=workspace_settings, cfg={"_password": ""}, stable_wait=0)

    assert processed == 1
    [output] = list(Path(workspace_settings["folders"]["output"]).glob("*.xlsx"))
    wb_out = load_workbook(output, read_only=True)
    try:
        assert wb_out.sheetnames == ["Export", "Import", "NOTE"]
        assert len(list(wb_out["Export"].iter_rows(values_only=True))) > 1
        assert len(list(wb_out["Import"].iter_rows(values_only=True))) > 1
    finally:
        wb_out.close()
    [log_row] = read_log_rows(workspace_settings)
    assert log_row["dataset"] == "both"
    assert int(log_row["row_count"]) > 0


def test_t67_scan_values_template_pipeline(monkeypatch, tmp_path, seeded_pg):
    patch_pg_connect(monkeypatch, seeded_pg)
    column_path = tmp_path / "column.yaml"
    template_path = tmp_path / "request_template.xlsx"
    column_path.write_text(yaml.safe_dump(column_cfg(columns_ready=False), sort_keys=False, allow_unicode=True), encoding="utf-8")

    runner.scan_columns(column_path=column_path, cfg={"_password": ""}, yes=True, today=runner.dt.date(2026, 6, 30))
    runner.scan_values(column_path=column_path, cfg={"_password": ""}, yes=True, today=runner.dt.date(2026, 6, 30))
    scanned = runner.load_yaml_file(column_path, {})
    export = scanned["datasets"]["export"]

    assert len(export["columns"]) == 32
    assert len(scanned["datasets"]["import"]["columns"]) == 44
    assert export["cardinality_cache"]["ma_nuoc"] == 5
    assert export["cardinality_cache"]["phuong_thuc_van_chuyen"] == 3
    assert set(export["value_cache"]["ma_nuoc"]) == {"CN", "KR", "JP", "VN", "US"}

    runner.make_request_template_v5(scanned, template_path)
    wb = load_workbook(template_path)
    try:
        assert wb.sheetnames == ["Request", "Cột Export", "Cột Import", "Values", "Tham chiếu"]
        assert wb["Values"]["A1"].value == "ma_nuoc"
        assert [wb["Values"].cell(row=row, column=1).value for row in range(2, 7)] == ["CN", "JP", "KR", "US", "VN"]
        assert defined_name_text(wb, "ma_nuoc_values") == "Values!$A$2:$A$6"
        ws = wb["Cột Export"]
        ma_nuoc_row = by_column_row(ws)["ma_nuoc"]
        dv = next(
            item
            for item in ws.data_validations.dataValidation
            if f"B{ma_nuoc_row}" in item.cells and item.type == "list"
        )
        assert dv.formula1 == "=ma_nuoc_values"
    finally:
        wb.close()
