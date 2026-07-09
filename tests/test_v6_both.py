import pytest
from openpyxl import load_workbook

import runner
from tests.test_v6_parse_vn import column_cfg


def make_both_request(tmp_path, export_values=None, import_values=None):
    cfg = column_cfg()
    path = tmp_path / "request_both.xlsx"
    runner.make_request_template_v5(cfg, path)
    wb = load_workbook(path)
    req = wb["Request"]
    req["B1"] = "Ana"
    req["B2"] = "both"
    req["B3"] = "2026"
    req["B4"] = "03"
    req["B7"] = "both-test"
    for sheet_name, values in (("Cột Export", export_values or {}), ("Cột Import", import_values or {})):
        ws = wb[sheet_name]
        by_col = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
        for column, cell_values in values.items():
            row = by_col[column]
            for excel_col, value in cell_values.items():
                ws[f"{excel_col}{row}"] = value
    wb.save(path)
    return path, cfg


def test_t53_both_parses_export_and_import_sheets(tmp_path):
    path, cfg = make_both_request(
        tmp_path,
        export_values={"ma_so": {"E": "8306", "H": "4"}},
        import_values={"out_im": {"I": "YES"}},
    )

    parsed = runner.parse_request_v6(path, cfg)

    assert parsed["bang"] == "both"
    assert parsed["export"]["filters"] == [{"col": "ma_so", "op": "prefix", "val": "8306", "digits": 4}]
    assert parsed["export"]["select_cols"] == ["ma_so"]
    assert parsed["import"]["filters"] == []
    assert parsed["import"]["select_cols"] == ["out_im"]


def test_t54_both_with_two_empty_sheets_rejects(tmp_path):
    path, cfg = make_both_request(tmp_path)

    with pytest.raises(runner.RequestError, match="Bảng=both nhưng cả 2 sheet đều trống"):
        runner.parse_request_v6(path, cfg)
