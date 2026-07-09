import pytest
from openpyxl import load_workbook

import runner
from tests.test_v5_e2e import make_legacy_v5_template


def column_cfg():
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": ["ma_so", "ma_nuoc", "out_01"],
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": ["ma_so", "out_im"],
            },
        },
        "operator_defaults": {"ma_so": "prefix"},
        "cardinality": {"threshold": 30, "sample_size": 1000, "skip_text_length": 100, "skip_columns": []},
    }


def make_v6_request(tmp_path, rows):
    cfg = column_cfg()
    path = tmp_path / "request_v6.xlsx"
    runner.make_request_template_v5(cfg, path)
    wb = load_workbook(path)
    req = wb["Request"]
    req["B1"] = "Ana"
    req["B2"] = "export"
    req["B3"] = "2026"
    req["B4"] = "03"
    req["B7"] = "parse-test"
    ws = wb["Cột Export"]
    by_col = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
    for column, values in rows.items():
        row = by_col[column]
        for excel_col, value in values.items():
            ws[f"{excel_col}{row}"] = value
    wb.save(path)
    return path, cfg


def test_t48_parse_prefix_cell_with_digits(tmp_path):
    path, cfg = make_v6_request(tmp_path, {"ma_so": {"E": "8306, 8307", "H": "4"}})

    parsed = runner.parse_request_v6(path, cfg)

    assert parsed["filters"] == [{"col": "ma_so", "op": "prefix", "val": "8306, 8307", "digits": 4}]
    assert parsed["select_cols"] == ["ma_so"]


def test_t49_v5_template_falls_back_with_warning(tmp_path):
    cfg = column_cfg()
    path = tmp_path / "request_v5.xlsx"
    make_legacy_v5_template(cfg, path)
    wb = load_workbook(path)
    req = wb["Request"]
    req["B1"] = "Ana"
    req["B2"] = "export"
    req["B3"] = "2026"
    req["B4"] = "03"
    req["B7"] = "legacy"
    ws = wb["Cột Export"]
    ws["B2"] = "prefix"
    ws["C2"] = "8306"
    wb.save(path)

    parsed = runner.parse_request_v6(path, cfg)

    assert parsed["filters"] == [{"col": "ma_so", "op": "prefix", "val": "8306"}]
    assert "template v5, khuyến khích v6" in parsed["warnings"]


def test_t50_digits_on_non_digits_operator_warns_and_ignores(tmp_path):
    path, cfg = make_v6_request(tmp_path, {"ma_nuoc": {"B": "CN", "H": "4"}})

    parsed = runner.parse_request_v6(path, cfg)

    assert parsed["filters"] == [{"col": "ma_nuoc", "op": "eq", "val": "CN", "digits": None}]
    assert any("op active không hỗ trợ Digits" in warning for warning in parsed["warnings"])


def test_t51_prefix_value_length_must_match_digits(tmp_path):
    path, cfg = make_v6_request(tmp_path, {"ma_so": {"E": "84", "H": "4"}})

    with pytest.raises(runner.RequestError, match="value '84' có 2 ký tự, Digits yêu cầu 4"):
        runner.parse_request_v6(path, cfg)


def test_t52_multi_op_same_row_rejects_suffix_with_wrong_digits(tmp_path):
    path, cfg = make_v6_request(tmp_path, {"ma_so": {"E": "8306, 8307", "G": "00", "H": "4"}})

    with pytest.raises(runner.RequestError, match="value '00' có 2 ký tự, Digits yêu cầu 4"):
        runner.parse_request_v6(path, cfg)


def test_t52b_multi_op_same_row_prefix_suffix_pass(tmp_path):
    path, cfg = make_v6_request(tmp_path, {"ma_so": {"E": "8306, 8307", "G": "0000", "H": "4"}})

    parsed = runner.parse_request_v6(path, cfg)

    assert parsed["filters"] == [
        {"col": "ma_so", "op": "prefix", "val": "8306, 8307", "digits": 4},
        {"col": "ma_so", "op": "suffix", "val": "0000", "digits": 4},
    ]
    assert parsed["select_cols"] == ["ma_so"]
