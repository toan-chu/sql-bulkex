import pytest
from openpyxl import load_workbook

pgserver = pytest.importorskip("pgserver")
psycopg2 = pytest.importorskip("psycopg2")

import runner
from tests.test_v5_e2e import fill_request, initial_column_cfg


@pytest.fixture
def pg_uri(tmp_path):
    if not hasattr(pgserver, "get_server"):
        pytest.skip("pgserver.get_server API is unavailable")
    pg = pgserver.get_server(str(tmp_path / "pgdata"), cleanup_mode="stop")
    try:
        yield pg.get_uri()
    finally:
        pg.cleanup()


def test_runner_v5_request_end_to_end_pgserver(monkeypatch, tmp_path, pg_uri):
    columns = ["ma_so_hang_hoa", "ma_nuoc_xuat_xu", "out_01"]
    conn = psycopg2.connect(pg_uri)
    conn.autocommit = True
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE public.x_y2025_01 (
                ma_so_hang_hoa text,
                ma_nuoc_xuat_xu text,
                out_01 text
            )
            """
        )
        cur.execute(
            "INSERT INTO public.x_y2025_01 VALUES (%s, %s, %s), (%s, %s, %s)",
            ("843610", "CN", "match", "0101", "VN", "skip"),
        )
    conn.close()

    template = tmp_path / "request_template.xlsx"
    requests = tmp_path / "requests"
    requests.mkdir()
    cfg = initial_column_cfg(columns=columns)
    cfg["datasets"]["export"]["schema"] = "public"
    runner.make_request_template_v5(cfg, template)
    fill_request(
        template,
        requests / "request.xlsx",
        anchors=[
            ("ma_so_hang_hoa", "prefix", "8436"),
            ("ma_nuoc_xuat_xu", "eq", "CN"),
        ],
        outputs=["out_01"],
    )

    settings = {
        "input_dir": str(requests),
        "output_dir": str(tmp_path / "out"),
        "filename_pattern": "{ts}_{user}_{request}",
        "max_rows_auto": 300000,
        "max_rows_hard": 3000000,
    }
    monkeypatch.setattr(runner, "load_column_config", lambda: cfg)
    monkeypatch.setattr(runner.portal, "connect", lambda cfg, dbname: psycopg2.connect(pg_uri))

    processed = runner.run_once(settings=settings, cfg={"_password": ""}, stable_wait=0)

    assert processed == 1
    outputs = list((tmp_path / "out").glob("*.xlsx"))
    assert len(outputs) == 1
    wb = load_workbook(outputs[0], read_only=True)
    assert "Data" in wb.sheetnames
    assert "NOTE" in wb.sheetnames
    rows = list(wb["Data"].iter_rows(values_only=True))
    assert rows[0] == ("ma_so_hang_hoa", "ma_nuoc_xuat_xu", "out_01")
    assert rows[1] == ("843610", "CN", "match")
    assert len(rows) == 2
    assert (requests / "processed").exists()
