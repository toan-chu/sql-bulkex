import pytest
from openpyxl import Workbook

import runner


EXPORT_COLUMNS = [
    "so_to_khai",
    "ma_so_hang_hoa",
    "ma_nuoc_xuat_xu",
    "tri_gia_usd",
    "ten_nguoi_xuat_khau",
    "ma_nguoi_nhap_khau",
] + [f"col_{i:02d}" for i in range(1, 28)]


def make_column_sheet(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cột Export"
    ws.append(["Cột", "Toán tử", "Giá trị", "Lấy về?"])
    for row in rows:
        ws.append(row)
    return ws


def parse_rows(rows, valid_cols=None, defaults=None):
    return runner.parse_column_sheet(
        make_column_sheet(rows),
        set(valid_cols or EXPORT_COLUMNS),
        defaults or {},
    )


def make_request_workbook(path, dataset="export", export_rows=None, import_rows=None):
    wb = Workbook()
    request = wb.active
    request.title = "Request"
    request_rows = [
        ("Người yêu cầu", "Ana"),
        ("Bảng", dataset),
        ("Năm", "2026"),
        ("Tháng", "03"),
        ("Tách file theo", ""),
        ("Xác nhận lớn", ""),
        ("Ghi chú / tên request", "daily"),
    ]
    for row in request_rows:
        request.append(row)

    export = wb.create_sheet("Cột Export")
    export.append(["Cột", "Toán tử", "Giá trị", "Lấy về?"])
    for row in export_rows or []:
        export.append(row)

    imp = wb.create_sheet("Cột Import")
    imp.append(["Cột", "Toán tử", "Giá trị", "Lấy về?"])
    for row in import_rows or []:
        imp.append(row)

    wb.save(path)
    return path


def column_cfg():
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": EXPORT_COLUMNS,
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": ["so_to_khai", "ma_so_hang_hoa"],
            },
        },
        "operator_defaults": {"ma_so_hang_hoa": "prefix"},
    }


def test_t4_prefix_filter_auto_selects_anchor_column():
    filters, select_cols, warnings = parse_rows([("ma_so_hang_hoa", "prefix", "8436", "")])

    assert filters == [{"col": "ma_so_hang_hoa", "op": "prefix", "val": "8436"}]
    assert select_cols == ["ma_so_hang_hoa"]
    assert warnings == []


def test_t5_blank_operator_uses_default_and_warns():
    filters, select_cols, warnings = parse_rows(
        [("ma_so_hang_hoa", "", "8436", "")],
        defaults={"ma_so_hang_hoa": "prefix"},
    )

    assert filters == [{"col": "ma_so_hang_hoa", "op": "prefix", "val": "8436"}]
    assert select_cols == ["ma_so_hang_hoa"]
    assert any("auto prefix" in warning for warning in warnings)


def test_t6_blank_operator_without_default_ignores_value_with_warning():
    filters, select_cols, warnings = parse_rows([("ma_so_hang_hoa", "", "8436", "")])

    assert filters == []
    assert select_cols == []
    assert any("Giá trị bỏ qua" in warning for warning in warnings)


def test_t7_operator_without_value_is_error():
    with pytest.raises(runner.RequestError, match="thiếu Giá trị"):
        parse_rows([("ma_so_hang_hoa", "eq", "", "")])


def test_t8_between_requires_two_values():
    with pytest.raises(runner.RequestError, match="cần đúng 2 giá trị"):
        parse_rows([("tri_gia_usd", "between", "1000", "")])


def test_t9_between_two_values_is_valid():
    filters, select_cols, warnings = parse_rows([("tri_gia_usd", "between", "1000,5000", "")])

    assert filters == [{"col": "tri_gia_usd", "op": "between", "val": "1000,5000"}]
    assert select_cols == ["tri_gia_usd"]
    assert warnings == []


def test_t10_between_rejects_three_values():
    with pytest.raises(runner.RequestError, match="cần đúng 2 giá trị"):
        parse_rows([("tri_gia_usd", "between", "1000,5000,9000", "")])


def test_t11_yes_without_filter_is_output_only():
    filters, select_cols, warnings = parse_rows([("so_to_khai", "", "", "YES")])

    assert filters == []
    assert select_cols == ["so_to_khai"]
    assert warnings == []


def test_t12_filter_overrides_no_and_still_selects_column():
    filters, select_cols, warnings = parse_rows([("ma_nuoc_xuat_xu", "eq", "CN", "NO")])

    assert filters == [{"col": "ma_nuoc_xuat_xu", "op": "eq", "val": "CN"}]
    assert select_cols == ["ma_nuoc_xuat_xu"]
    assert warnings == []


def test_t13_two_anchors_and_three_outputs_select_five_columns():
    rows = [
        ("ma_so_hang_hoa", "prefix", "8436", ""),
        ("ma_nuoc_xuat_xu", "eq", "CN", ""),
        ("so_to_khai", "", "", "YES"),
        ("ten_nguoi_xuat_khau", "", "", "YES"),
        ("ma_nguoi_nhap_khau", "", "", "YES"),
    ]
    rows.extend((f"col_{i:02d}", "", "", "") for i in range(1, 28))

    filters, select_cols, warnings = parse_rows(rows)

    assert len(filters) == 2
    assert len(select_cols) == 5
    assert warnings == []


def test_t14_empty_request_has_no_filters_or_outputs(tmp_path):
    path = make_request_workbook(tmp_path / "request.xlsx")

    with pytest.raises(runner.RequestError, match="Chưa chọn cột filter"):
        runner.parse_request_v5(path, column_cfg())


def test_t15_export_request_reads_export_sheet_not_import_sheet(tmp_path):
    path = make_request_workbook(
        tmp_path / "request.xlsx",
        dataset="export",
        export_rows=[("so_to_khai", "", "", "YES")],
        import_rows=[("ma_so_hang_hoa", "starts", "8436", "")],
    )

    parsed = runner.parse_request_v5(path, column_cfg())

    assert parsed["dataset_name"] == "export"
    assert parsed["select_cols"] == ["so_to_khai"]
    assert parsed["filters"] == []


def test_t16_invalid_column_in_sheet_warns_and_skips():
    filters, select_cols, warnings = parse_rows([("db_drift_col", "eq", "CN", "")])

    assert filters == []
    assert select_cols == []
    assert warnings == ["Cột không hợp lệ trong sheet: db_drift_col"]


def test_t17_invalid_operator_is_error():
    with pytest.raises(runner.RequestError, match="toán tử không hợp lệ"):
        parse_rows([("ma_so_hang_hoa", "starts", "8436", "")])
