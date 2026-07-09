import builtins
import csv
from pathlib import Path

from openpyxl import Workbook

import runner


def log_settings(tmp_path):
    pending = tmp_path / "01_Pending"
    approved = tmp_path / "02_Approved"
    output = tmp_path / "03_Output"
    log_dir = tmp_path / "log"
    pending.mkdir()
    approved.mkdir()
    output.mkdir()
    log_dir.mkdir()
    return {
        "folders": {
            "pending": str(pending),
            "approved": str(approved),
            "output": str(output),
        },
        "log": {
            "requests_csv": str(log_dir / "requests.csv"),
            "runner_log": str(log_dir / "runner.log"),
            "portal_log": str(log_dir / "portal.log"),
        },
        "filename_pattern": "{ts}_{user}_{request}",
        "max_rows_auto": 300000,
        "max_rows_hard": 3000000,
    }


def write_request(path, requester="Hoa", last_modified_by="VSTREAM\\hoa.nguyen"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Request"
    ws["A1"] = "Người yêu cầu"
    ws["B1"] = requester
    wb.properties.lastModifiedBy = last_modified_by
    wb.save(path)
    return path


def read_rows(settings):
    with open(settings["log"]["requests_csv"], newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def test_t62_success_appends_structured_request_log(monkeypatch, tmp_path):
    settings = log_settings(tmp_path)
    approved = Path(settings["folders"]["approved"])
    output = Path(settings["folders"]["output"])
    request = write_request(approved / "request_x.xlsx")

    def fake_process(path, cfg, conns, settings):
        out = output / "result.xlsx"
        out.write_text("ok", encoding="utf-8")
        runner.mark_done(path)
        return {"dataset": "export", "row_count": 12, "output_file": out.name}

    monkeypatch.setattr(runner, "process_request_file", fake_process)

    processed = runner.run_once(settings=settings, cfg={}, stable_wait=0)

    assert processed == 1
    rows = read_rows(settings)
    assert len(rows) == 1
    row = rows[0]
    assert list(row.keys()) == runner.REQUEST_LOG_HEADER
    assert row["status"] == "success"
    assert row["row_count"] == "12"
    assert float(row["duration_sec"]) >= 0
    assert row["output_file"] == "result.xlsx"
    assert row["file_name"] == request.name


def test_t63_rejected_request_logs_error(monkeypatch, tmp_path):
    settings = log_settings(tmp_path)
    approved = Path(settings["folders"]["approved"])
    write_request(approved / "request_x.xlsx")

    def fake_process(path, cfg, conns, settings):
        raise runner.RequestError("Bảng không hợp lệ: bad")

    monkeypatch.setattr(runner, "process_request_file", fake_process)

    processed = runner.run_once(settings=settings, cfg={}, stable_wait=0)

    assert processed == 0
    rows = read_rows(settings)
    assert len(rows) == 1
    row = rows[0]
    assert row["status"] == "rejected"
    assert row["row_count"] == "0"
    assert float(row["duration_sec"]) >= 0
    assert row["error"] == "Bảng không hợp lệ: bad"


def test_t64_log_request_csv_creates_header_when_missing(tmp_path):
    settings = log_settings(tmp_path)

    runner.log_request_csv(
        settings,
        1,
        requester_cell="Hoa",
        requester_meta="VSTREAM\\hoa.nguyen",
        file_name="request.xlsx",
        dataset="export",
        row_count=1,
        duration_sec=0.25,
        status="success",
        output_file="out.xlsx",
    )

    with open(settings["log"]["requests_csv"], newline="", encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    assert header == runner.REQUEST_LOG_HEADER


def test_t64b_permission_error_retries_then_logs_warning(monkeypatch, tmp_path):
    settings = log_settings(tmp_path)
    logs = []
    attempts = []
    real_open = builtins.open

    def locked_open(path, *args, **kwargs):
        if str(path) == settings["log"]["requests_csv"]:
            attempts.append(path)
            raise PermissionError("locked by Excel")
        return real_open(path, *args, **kwargs)

    monkeypatch.setattr(builtins, "open", locked_open)
    monkeypatch.setattr(runner.time, "sleep", lambda seconds: None)
    monkeypatch.setattr(runner, "log_event", logs.append)

    ok = runner.log_request_csv(settings, 1, file_name="request.xlsx", status="success")

    assert ok is False
    assert len(attempts) == 3
    assert any("[REQUEST_LOG] Không ghi được" in item and "locked by Excel" in item for item in logs)


def test_t64c_requester_cell_and_last_modified_by_are_logged(monkeypatch, tmp_path):
    settings = log_settings(tmp_path)
    approved = Path(settings["folders"]["approved"])
    write_request(approved / "request_x.xlsx", requester="Hoa", last_modified_by="VSTREAM\\hoa.nguyen")

    def fake_process(path, cfg, conns, settings):
        runner.mark_done(path)
        return {"dataset": "import", "row_count": 3, "output_file": "out.xlsx"}

    monkeypatch.setattr(runner, "process_request_file", fake_process)

    runner.run_once(settings=settings, cfg={}, stable_wait=0)

    [row] = read_rows(settings)
    assert row["requester_cell"] == "Hoa"
    assert row["requester_meta"] == "VSTREAM\\hoa.nguyen"
