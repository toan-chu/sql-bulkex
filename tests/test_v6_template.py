from openpyxl import load_workbook

import runner


EXPORT_COLUMNS = ["ma_nuoc", "ma_so_hang_hoa", "mo_ta_hang_hoa"]
IMPORT_COLUMNS = ["ma_nuoc", "ma_nguoi_nhap_khau"]
VN_OPERATOR_LABELS = ["Bằng", "Trong danh sách", "Trong khoảng", "Bắt đầu bằng", "Chứa", "Kết thúc bằng"]


def column_cfg():
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": EXPORT_COLUMNS,
                "cardinality_cache": {"ma_nuoc": 4},
                "value_cache": {"ma_nuoc": ["CN", "VN", "KR", "JP"]},
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": IMPORT_COLUMNS,
                "cardinality_cache": {},
                "value_cache": {},
            },
        },
        "operator_defaults": {},
        "cardinality": {
            "threshold": 30,
            "sample_size": 1000,
            "skip_text_length": 100,
            "skip_columns": [],
        },
    }


def build_template(tmp_path):
    output = tmp_path / "request_template.xlsx"
    runner.make_request_template_v5(column_cfg(), output)
    return load_workbook(output)


def validation_for_cell(ws, cell_ref, validation_type=None):
    for dv in ws.data_validations.dataValidation:
        if cell_ref in dv.cells and (validation_type is None or dv.type == validation_type):
            return dv
    return None


def defined_name_text(wb, name):
    defined = wb.defined_names.get(name)
    if isinstance(defined, list):
        defined = defined[0]
    return defined.attr_text if defined is not None else None


def conditional_rules(ws, cell_range):
    for item in ws.conditional_formatting:
        if cell_range in str(item):
            return ws.conditional_formatting[item]
    return []


def test_t42_make_template_has_five_v6_sheets_with_hidden_values(tmp_path):
    wb = build_template(tmp_path)

    assert wb.sheetnames == ["Request", "Cột Export", "Cột Import", "Values", "Tham chiếu"]
    assert wb["Values"].sheet_state == "hidden"
    assert wb["Request"]["A8"].value == "Người duyệt (Admin điền sau approve)"
    assert wb["Request"].print_area == "'Request'!$A$1:$C$8"


def test_t43_column_export_has_digits_header(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]

    header = [ws.cell(row=1, column=col).value for col in range(1, 10)]
    assert header == ["Cột"] + VN_OPERATOR_LABELS + ["Digits", "Lấy về?"]
    assert ws.column_dimensions["B"].width == 14
    assert ws.column_dimensions["G"].width == 14
    assert ws.column_dimensions["H"].width == 10


def test_t44_value_dropdown_only_applies_to_eq_and_in_cells(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]
    eq_dv = validation_for_cell(ws, "B2", "list")
    in_dv = validation_for_cell(ws, "C2", "list")

    assert eq_dv is not None
    assert eq_dv.formula1 == "=ma_nuoc_values"
    assert eq_dv.errorStyle == "information"
    assert in_dv is not None
    assert in_dv.formula1 == "=ma_nuoc_values"
    assert in_dv.errorStyle == "information"
    for cell in ("D2", "E2", "F2", "G2"):
        assert validation_for_cell(ws, cell, "list") is None


def test_t45_digits_column_has_gray_out_conditional_formatting(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]
    rules = conditional_rules(ws, "H2:H4")

    assert any(
        rule.type == "expression"
        and "AND(ISBLANK($E2), ISBLANK($G2))" in "".join(rule.formula)
        and rule.dxf.fill.fgColor.rgb == "00E7E6E6"
        and rule.dxf.font.color.rgb == "00A6A6A6"
        for rule in rules
    )


def test_t46_values_sheet_hidden_and_named_range_for_low_cardinality_column(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Values"]

    assert ws.sheet_state == "hidden"
    assert ws["A1"].value == "ma_nuoc"
    assert [ws.cell(row=row, column=1).value for row in range(2, 6)] == ["CN", "VN", "KR", "JP"]
    assert defined_name_text(wb, "ma_nuoc_values") == "Values!$A$2:$A$5"


def test_t47_value_cell_for_cached_column_references_named_range(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]
    dv = validation_for_cell(ws, "B2", "list")

    assert dv is not None
    assert dv.formula1 == "=ma_nuoc_values"


def test_t47b_digits_cell_has_common_values_dropdown(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]
    dv = validation_for_cell(ws, "H2", "list")

    assert dv is not None
    assert dv.formula1 == '"2, 4, 6, 8, 10, 13"'
    assert dv.allow_blank
    assert dv.errorStyle == "information"


def test_t47c_row_anchor_conditional_formatting_counts_operator_cells(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Cột Export"]
    rules = conditional_rules(ws, "A2:I4")

    assert any(
        rule.type == "expression"
        and "COUNTA($B2:$G2)>0" in "".join(rule.formula)
        and rule.dxf.fill.fgColor.rgb == "00FFF2CC"
        for rule in rules
    )


def test_requester_dropdown_uses_reference_named_range(tmp_path):
    wb = build_template(tmp_path)
    request = wb["Request"]
    reference = wb["Tham chiếu"]
    dv = validation_for_cell(request, "B1", "list")
    header_row = next(
        row
        for row in range(1, reference.max_row + 1)
        if reference.cell(row=row, column=1).value == "Danh sách người yêu cầu (Admin điền)"
    )

    assert defined_name_text(wb, "nguoi_yeu_cau_list") == f"'Tham chiếu'!$A${header_row + 1}:$A${header_row + 30}"
    assert dv is not None
    assert dv.formula1 == "=nguoi_yeu_cau_list"
    assert dv.allow_blank
    assert dv.errorStyle == "information"


def test_request_sheet_has_guidance_column(tmp_path):
    wb = build_template(tmp_path)
    ws = wb["Request"]
    hints = [ws.cell(row=row, column=3).value for row in range(1, 9)]

    assert all(hints)
    assert "both" in ws["C2"].value
    assert ws.column_dimensions["C"].width == 45
