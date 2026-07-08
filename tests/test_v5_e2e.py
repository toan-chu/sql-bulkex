import shutil
from pathlib import Path

import yaml
from openpyxl import load_workbook, Workbook

import runner


EXPORT_COLUMNS = [
    "ma_so_hang_hoa",
    "ma_nuoc_xuat_xu",
] + [f"out_{i:02d}" for i in range(1, 21)] + [f"extra_{i:02d}" for i in range(1, 11)]


def initial_column_cfg(columns=None, defaults=None):
    return {
        "datasets": {
            "export": {
                "database": "vn_export",
                "schema": "vietnam_export",
                "tables": "x_y{year}_{month}",
                "columns": columns or [],
            },
            "import": {
                "database": "vn_import",
                "schema": "vietnam_import",
                "tables": "i_y{year}_{month}",
                "columns": ["imp_col"],
            },
        },
        "operator_defaults": defaults or {},
    }


def write_column_yaml(path, cfg):
    path.write_text(yaml.safe_dump(cfg, sort_keys=False, allow_unicode=True), encoding="utf-8")


class InfoSchemaCursor:
    def __init__(self, table_columns):
        self.table_columns = table_columns
        self.tables = set(table_columns)
        self.rows = []
        self.description = []
        self.closed = False

    def execute(self, query, params=None):
        query_text = str(query)
        if "information_schema.columns" in query_text:
            table = params[1]
            columns = self.table_columns.get(table, [])
            if "data_type" in query_text:
                self.rows = [(column, "text") for column in columns]
            else:
                self.rows = [(column,) for column in columns]
            return
        if "information_schema.tables" in query_text:
            table = params[1]
            self.rows = [(1,)] if table in self.tables else []
            return
        self.rows = []

    def fetchone(self):
        return self.rows[0] if self.rows else None

    def fetchall(self):
        return self.rows

    def close(self):
        self.closed = True


class FakeConn:
    closed = False

    def __init__(self, table_columns):
        self.table_columns = table_columns
        self.rollbacks = 0

    def set_client_encoding(self, encoding):
        self.encoding = encoding

    def cursor(self, name=None):
        return InfoSchemaCursor(self.table_columns)

    def rollback(self):
        self.rollbacks += 1

    def close(self):
        self.closed = True


def fill_request(template_path, request_path, anchors, outputs, values=None):
    shutil.copyfile(template_path, request_path)
    wb = load_workbook(request_path)
    req = wb["Request"]
    req["B1"] = "Ana"
    req["B2"] = "export"
    req["B3"] = "2025"
    req["B4"] = "01"
    req["B7"] = "daily"
    values = values or {}
    ws = wb["Cột Export"]
    by_col = {ws.cell(row=row, column=1).value: row for row in range(2, ws.max_row + 1)}
    output_col = next(
        col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value == "Lấy về?"
    )
    for column, op, value in anchors:
        row = by_col[column]
        if op:
            ws.cell(row=row, column=2, value=op)
        ws.cell(row=row, column=3, value=value)
    for column in outputs:
        ws.cell(row=by_col[column], column=output_col, value="YES")
    for column, value in values.items():
        ws.cell(row=by_col[column], column=3, value=value)
    wb.save(request_path)
    return request_path


def patch_v5_boundaries(monkeypatch, tmp_path, table_columns, row_count=2):
    last_state = {}
    original_make_jobs = runner.portal.make_jobs
    conn = FakeConn(table_columns)

    def make_jobs_capture(state, cur):
        last_state.clear()
        last_state.update(state)
        return original_make_jobs(state, cur)

    def fake_export(conn, query, params, headers, filepath, notes):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.append(headers)
        for index in range(row_count):
            ws.append([f"r{index}_{header}" for header in headers])
        note = wb.create_sheet("NOTE")
        for row in notes:
            note.append(list(row))
        wb.save(filepath)
        return row_count

    monkeypatch.setattr(runner.portal, "connect", lambda cfg, dbname: FakeConn(table_columns))
    monkeypatch.setattr(runner.portal, "get_conn", lambda conns, cfg, dbname: conn)
    monkeypatch.setattr(runner.portal, "make_jobs", make_jobs_capture)
    monkeypatch.setattr(runner.portal, "count_rows", lambda cur, query, params: row_count)
    monkeypatch.setattr(runner, "fetch_headers", lambda conn, query, params: list(last_state["cols"]))
    monkeypatch.setattr(runner, "export_xlsx_v5_with_note", fake_export)
    monkeypatch.setattr(runner, "TMP_DIR", tmp_path / "tmp")
    monkeypatch.setattr(runner, "LOG_DIR", tmp_path)
    monkeypatch.setattr(runner, "RUNNER_LOG_FILE", tmp_path / "runner.log")
    return last_state


def test_t20_v5_full_flow_scan_template_run_once(monkeypatch, tmp_path):
    column_path = tmp_path / "column.yaml"
    template_path = tmp_path / "request_template.xlsx"
    requests = tmp_path / "requests"
    output = tmp_path / "out"
    requests.mkdir()
    table_columns = {"x_y2025_01": EXPORT_COLUMNS}
    write_column_yaml(column_path, initial_column_cfg())
    monkeypatch.setattr(runner, "COLUMN_FILE", column_path)
    monkeypatch.setattr(runner, "REQUEST_TEMPLATE_FILE", template_path)
    last_state = patch_v5_boundaries(monkeypatch, tmp_path, table_columns, row_count=2)

    runner.scan_columns(column_path=column_path, cfg={"_password": "secret"}, yes=True)
    assert runner.main(["--make-template"]) == 0
    fill_request(
        template_path,
        requests / "request.xlsx",
        anchors=[
            ("ma_so_hang_hoa", "prefix", "8436"),
            ("ma_nuoc_xuat_xu", "eq", "CN"),
        ],
        outputs=[f"out_{i:02d}" for i in range(1, 21)],
    )

    processed = runner.run_once(
        settings={
            "input_dir": str(requests),
            "output_dir": str(output),
            "filename_pattern": "{ts}_{user}_{request}",
            "max_rows_auto": 300000,
            "max_rows_hard": 3000000,
        },
        cfg={"_password": "secret"},
        stable_wait=0,
    )

    assert processed == 1
    outputs = list(output.glob("*.xlsx"))
    assert len(outputs) == 1
    wb = load_workbook(outputs[0], read_only=True)
    rows = list(wb["Data"].iter_rows(values_only=True))
    assert len(rows[0]) == 22
    assert len(rows) == 3
    assert len(last_state["filters"]) == 2
    assert len(last_state["cols"]) == 22
    assert (requests / "processed").exists()


def test_t21_v5_complex_case_builds_10_filters_and_20_selects(monkeypatch, tmp_path):
    columns = [f"f_{i:02d}" for i in range(1, 11)] + [f"out_{i:02d}" for i in range(1, 11)]
    table_columns = {"x_y2025_01": columns}
    runner.make_request_template_v5(initial_column_cfg(columns=columns), tmp_path / "template.xlsx")
    request_path = fill_request(
        tmp_path / "template.xlsx",
        tmp_path / "request.xlsx",
        anchors=[(f"f_{i:02d}", "eq", f"v{i}") for i in range(1, 11)],
        outputs=[f"out_{i:02d}" for i in range(1, 11)],
    )
    parsed = runner.parse_request_v5(request_path, initial_column_cfg(columns=columns))
    captured = {}
    monkeypatch.setattr(runner.portal, "make_jobs", lambda state, cur: captured.update(state) or [("job", "db", "SQL", [])])

    runner.build_jobs_from_v5_request(parsed, InfoSchemaCursor(table_columns))

    assert len(captured["filters"]) == 10
    assert len(captured["cols"]) == 20


def test_t22_v5_auto_default_uses_prefix_like_param(tmp_path):
    columns = ["ma_so_hang_hoa", "out_01"]
    table_columns = {"x_y2025_01": columns}
    template = tmp_path / "template.xlsx"
    request = tmp_path / "request.xlsx"
    cfg = initial_column_cfg(columns=columns, defaults={"ma_so_hang_hoa": "prefix"})
    runner.make_request_template_v5(cfg, template)
    fill_request(
        template,
        request,
        anchors=[],
        outputs=["out_01"],
        values={"ma_so_hang_hoa": "8436"},
    )
    parsed = runner.parse_request_v5(request, cfg)

    _dbname, (jobs, notes, selected) = runner.build_jobs_from_v5_request(parsed, InfoSchemaCursor(table_columns))

    assert jobs[0][3] == ["8436%"]
    assert selected == ["ma_so_hang_hoa", "out_01"]
    assert ("WARNING", "Cột ma_so_hang_hoa: auto prefix (user không chọn toán tử)") in notes
