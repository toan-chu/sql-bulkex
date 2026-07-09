#!/usr/bin/env python3
"""Headless SQL BulkEx request runner."""

import argparse
import copy
import csv
import datetime as dt
import os
import shutil
import subprocess
import sys
import time
from pathlib import Path

sys.dont_write_bytecode = True

import psycopg2
from psycopg2 import sql
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation

import portal
from operators import OperatorBuilder, OperatorRegistryError, OperatorValueError


BASE_DIR = Path(__file__).resolve().parent
SETTINGS_FILE = BASE_DIR / "settings.yaml"
COLUMN_FILE = BASE_DIR / "column.yaml"
OPERATORS_FILE = BASE_DIR / "operators.yaml"
REQUEST_TEMPLATE_FILE = BASE_DIR / "request_template.xlsx"
LOG_DIR = BASE_DIR / "log"
TMP_DIR = LOG_DIR / "tmp"
RUNNER_LOG_FILE = LOG_DIR / "runner.log"
XLSX_ROW_LIMIT = 1_000_000
REJECT_PREFIX = "[LOI]_"
DONE_PREFIX = "[DONE] "
DONE_TIMESTAMP_PREFIX = "[DONE "
REQUEST_V5_LABELS_ORDER = [
    "Người yêu cầu",
    "Bảng",
    "Năm",
    "Tháng",
    "Tách file theo",
    "Xác nhận lớn",
    "Ghi chú / tên request",
]
REQUEST_V6_LABELS_ORDER = REQUEST_V5_LABELS_ORDER + ["Người duyệt (Admin điền sau approve)"]
COLUMN_SCAN_SKELETON = {
    "datasets": {
        "export": {
            "database": "",
            "schema": "",
            "tables": "",
            "columns": [],
            "cardinality_cache": {},
            "value_cache": {},
        }
    },
    "operator_defaults": {},
    "cardinality": {
        "threshold": 30,
        "sample_size": 1000,
        "skip_text_length": 100,
        "skip_columns": [],
    },
}
DEFAULT_CARDINALITY = {
    "threshold": 30,
    "sample_size": 1000,
    "skip_text_length": 100,
    "skip_columns": [],
}
COLUMN_SETUP_MESSAGE = (
    "Chưa điền datasets trong column.yaml. "
    "Vui lòng điền database/schema/tables rồi chạy lại."
)
REQUEST_V5_LABELS = {
    "Người yêu cầu": "user",
    "Bảng": "bang",
    "Năm": "year",
    "Tháng": "month",
    "Tách file theo": "split",
    "Xác nhận lớn": "large_confirm",
    "Ghi chú / tên request": "request_name",
}

DEFAULT_SETTINGS = {
    "folders": {
        "pending": str(BASE_DIR / "SQL-BulkEx-Workspace" / "01_Pending"),
        "approved": str(BASE_DIR / "SQL-BulkEx-Workspace" / "02_Approved"),
        "output": str(BASE_DIR / "SQL-BulkEx-Workspace" / "03_Output"),
    },
    "poll_seconds": 120,
    "filename_pattern": "{ts}_{user}_{request}",
    "max_rows_auto": 300000,
    "max_rows_hard": 3000000,
    "onedrive_freeup": {
        "enabled": True,
        "approved_delay_hours": 2,
        "output_delay_days": 7,
    },
    "log": {
        "requests_csv": str(BASE_DIR / "log" / "requests.csv"),
        "runner_log": str(RUNNER_LOG_FILE),
        "portal_log": str(LOG_DIR / "portal.log"),
    },
}
LEGACY_DEFAULT_SETTINGS = {
    "input_dir": str(BASE_DIR / "requests"),
    "output_dir": str(BASE_DIR / "exports"),
    "poll_seconds": 120,
    "filename_pattern": "{ts}_{user}_{request}",
    "max_rows_auto": 300000,
    "max_rows_hard": 3000000,
}
V5_SETTINGS_WARNING = (
    "WARNING deprecation: settings.yaml đang dùng schema v5, "
    "migrate sang folders.pending/approved/output"
)


class RequestError(Exception):
    """A user-fixable request file problem."""


class RunnerConfigError(Exception):
    """A runner startup config problem."""


def configure_stdio():
    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8", errors="replace")


def log_event(message):
    LOG_DIR.mkdir(exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(RUNNER_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(f"[{stamp}] {message}\n")


def load_yaml_file(path, default):
    path = Path(path)
    if not path.exists():
        return dict(default)
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    if isinstance(default, dict):
        merged = dict(default)
        merged.update(data)
        return merged
    return data


def load_settings(path=None):
    path = Path(path or SETTINGS_FILE)
    if not path.exists():
        return dict(DEFAULT_SETTINGS)
    with open(path, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    if "folders" not in data and "input_dir" in data:
        merged = dict(LEGACY_DEFAULT_SETTINGS)
        merged.update(data)
        merged["_deprecated_v5_schema"] = True
        return merged
    merged = dict(DEFAULT_SETTINGS)
    merged.update(data)
    folders = dict(DEFAULT_SETTINGS["folders"])
    folders.update(data.get("folders") or {})
    merged["folders"] = folders
    onedrive_freeup = dict(DEFAULT_SETTINGS["onedrive_freeup"])
    onedrive_freeup.update(data.get("onedrive_freeup") or {})
    merged["onedrive_freeup"] = onedrive_freeup
    log_cfg = dict(DEFAULT_SETTINGS["log"])
    log_cfg.update(data.get("log") or {})
    merged["log"] = log_cfg
    return merged


def write_yaml_file(path, data):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        yaml.safe_dump(data, f, sort_keys=False, allow_unicode=True)


def create_column_skeleton(path=None):
    path = path or COLUMN_FILE
    write_yaml_file(path, COLUMN_SCAN_SKELETON)
    return Path(path)


def load_column_config(path=None):
    path = path or COLUMN_FILE
    data = load_yaml_file(path, {})
    datasets = data.get("datasets") or {}
    if not datasets:
        raise RunnerConfigError("column.yaml chưa có datasets. Chạy: python runner.py --scan-columns")
    for name, dataset in datasets.items():
        for key in ("database", "schema", "tables", "columns"):
            if not dataset.get(key):
                raise RunnerConfigError(f"Dataset {name} thiếu field: {key}")
    op_defaults = data.get("operator_defaults") or {}
    valid_ops = load_operator_builder().valid_keys()
    for column, op in op_defaults.items():
        if op not in valid_ops:
            raise RunnerConfigError(f"operator_defaults.{column} không hợp lệ: {op}")
    validate_cardinality_config(data.get("cardinality") or {})
    return data


def validate_cardinality_config(config):
    threshold = int(config.get("threshold", DEFAULT_CARDINALITY["threshold"]))
    sample_size = int(config.get("sample_size", DEFAULT_CARDINALITY["sample_size"]))
    if threshold < 1 or threshold > 500:
        raise RunnerConfigError("cardinality.threshold phải trong khoảng 1-500")
    if sample_size < 100 or sample_size > 10000:
        raise RunnerConfigError("cardinality.sample_size phải trong khoảng 100-10000")


def cardinality_settings(column_cfg):
    config = dict(DEFAULT_CARDINALITY)
    config.update(column_cfg.get("cardinality") or {})
    validate_cardinality_config(config)
    config["threshold"] = int(config["threshold"])
    config["sample_size"] = int(config["sample_size"])
    config["skip_text_length"] = int(config["skip_text_length"])
    config["skip_columns"] = list(config.get("skip_columns") or [])
    return config


def ensure_cardinality_schema(column_cfg):
    config = dict(DEFAULT_CARDINALITY)
    config.update(column_cfg.get("cardinality") or {})
    column_cfg["cardinality"] = config
    for dataset in (column_cfg.get("datasets") or {}).values():
        dataset.setdefault("cardinality_cache", {})
        dataset.setdefault("value_cache", {})
    return column_cfg


def load_operator_builder(path=None):
    path = path or OPERATORS_FILE
    try:
        return OperatorBuilder(path)
    except OperatorRegistryError as e:
        raise RunnerConfigError(str(e)) from e


def load_connection_config():
    cfg = portal.load_config()
    password = portal.config_or_file_password(cfg)
    if not password:
        raise RunnerConfigError(
            "Runner cần password trong connection.yaml hoặc file .password. "
            "Tạo file .password cạnh runner.py vì runner chạy headless/pythonw không thể hỏi getpass."
        )
    cfg["_password"] = password
    return cfg


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_sheet_request(sheet):
    values = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        label = cell_text(row[0])
        if not label:
            continue
        key = REQUEST_V5_LABELS.get(label)
        if key:
            values[key] = cell_text(row[1])
    return values


def validate_op_value(col, op, val, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    try:
        op_builder.validate(col, op, val)
    except OperatorValueError as e:
        raise RequestError(str(e)) from e


def parse_column_sheet(sheet, valid_cols, op_defaults, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    valid_cols = set(valid_cols)
    op_defaults = op_defaults or {}
    filters = []
    select_cols = []
    warnings = []

    max_col = 5 if (sheet.max_column or 0) >= 5 else 4
    for row in sheet.iter_rows(min_row=2, max_col=max_col, values_only=True):
        values = [cell_text(cell) for cell in row]
        if max_col >= 5:
            col, op, val, _digits, out = values
        else:
            col, op, val, out = values
        if not col:
            continue
        if col not in valid_cols:
            warnings.append(f"Cột không hợp lệ trong sheet: {col}")
            continue

        op = op_builder.normalize_operator(op) if op else ""
        out = out.upper() if out else ""
        has_op = op in op_builder.valid_keys()
        has_val = bool(val)

        if op and not has_op:
            valid = ", ".join(op_builder.display_order)
            raise RequestError(f"Cột {col}: toán tử không hợp lệ '{op}'. Chỉ chấp nhận: {valid}")

        if has_op and has_val:
            validate_op_value(col, op, val, op_builder)
            filters.append({"col": col, "op": op, "val": val})
            if col not in select_cols:
                select_cols.append(col)
        elif has_op and not has_val:
            raise RequestError(f"Cột {col}: có toán tử '{op}' nhưng thiếu Giá trị.")
        elif not has_op and has_val:
            default_op = op_defaults.get(col)
            if default_op:
                default_op = op_builder.normalize_operator(default_op)
            if default_op and default_op in op_builder.valid_keys():
                validate_op_value(col, default_op, val, op_builder)
                filters.append({"col": col, "op": default_op, "val": val})
                if col not in select_cols:
                    select_cols.append(col)
                warnings.append(f"Cột {col}: auto {default_op} (user không chọn toán tử)")
            else:
                warnings.append(
                    f"Cột {col}: có Giá trị nhưng thiếu Toán tử và không có default. Giá trị bỏ qua."
                )
        elif out == "YES" and col not in select_cols:
            select_cols.append(col)

    return filters, select_cols, warnings


def is_v6_column_sheet(sheet, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    header = [
        cell_text(cell)
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, max_col=9, values_only=True), ())
    ]
    return len(header) >= 2 and header[1] == op_builder.operators[op_builder.display_order[0]]["display"]


def parse_column_sheet_v6_multi_op(sheet, valid_cols, op_defaults, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    if not is_v6_column_sheet(sheet, op_builder):
        filters, select_cols, warnings = parse_column_sheet(sheet, valid_cols, op_defaults, op_builder)
        warnings.append("template v5, khuyến khích v6")
        return filters, select_cols, warnings

    valid_cols = set(valid_cols)
    filters = []
    select_cols = []
    warnings = []
    op_keys_in_order = list(op_builder.display_order)
    if len(op_keys_in_order) != 6:
        raise RequestError("operators.yaml display_order phải có đúng 6 toán tử cho template v6.")

    for row in sheet.iter_rows(min_row=2, max_col=9, values_only=True):
        cells = [cell_text(cell) for cell in row]
        col = cells[0]
        if not col:
            continue
        if col not in valid_cols:
            warnings.append(f"Cột không hợp lệ trong sheet: {col}")
            continue

        op_values = {op_keys_in_order[i]: cells[i + 1] for i in range(6)}
        digits_raw = cells[7]
        out = cells[8].upper() if cells[8] else ""
        try:
            digits_int = op_builder.normalize_digits(digits_raw) if digits_raw else None
        except OperatorValueError as e:
            raise RequestError(str(e)) from e

        active_ops = {op: val for op, val in op_values.items() if val}
        if not active_ops:
            if digits_int is not None:
                warnings.append(f"Cột {col}: có Digits nhưng không có op nào active, bỏ qua Digits")
            if out == "YES" and col not in select_cols:
                select_cols.append(col)
            continue

        digits_used = False
        for op, val in active_ops.items():
            spec = op_builder.operators[op]
            digits_for_op = digits_int if spec.get("supports_digits") else None
            try:
                op_builder.validate(col, op, val, digits_for_op)
            except OperatorValueError as e:
                raise RequestError(str(e)) from e
            filters.append({"col": col, "op": op, "val": val, "digits": digits_for_op})
            if digits_for_op is not None:
                digits_used = True
            if col not in select_cols:
                select_cols.append(col)

        if digits_int is not None and not digits_used:
            warnings.append(f"Cột {col}: có Digits nhưng op active không hỗ trợ Digits, bỏ qua")

    return filters, select_cols, warnings


def parse_request_v5(path, column_cfg):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Request" not in wb.sheetnames:
            raise RequestError("Thiếu sheet Request.")
        req = parse_sheet_request(wb["Request"])
        dataset_name = cell_text(req.get("bang")).lower()
        datasets = column_cfg.get("datasets") or {}
        if dataset_name not in datasets:
            raise RequestError(f"Bảng không hợp lệ: {req.get('bang', '')}")

        dataset = datasets[dataset_name]
        valid_cols = set(dataset.get("columns") or [])
        col_sheet_name = "Cột Export" if dataset_name == "export" else "Cột Import"
        if col_sheet_name not in wb.sheetnames:
            raise RequestError(f"Thiếu sheet {col_sheet_name}.")
        filters, select_cols, warnings = parse_column_sheet(
            wb[col_sheet_name],
            valid_cols,
            column_cfg.get("operator_defaults") or {},
        )
        if not filters and not select_cols:
            raise RequestError("Chưa chọn cột filter cũng chưa chọn cột lấy về.")

        return {
            "request": req,
            "dataset": dataset,
            "dataset_name": dataset_name,
            "filters": filters,
            "select_cols": select_cols,
            "warnings": warnings,
        }
    finally:
        wb.close()


def parse_request_v6(path, column_cfg, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        if "Request" not in wb.sheetnames:
            raise RequestError("Thiếu sheet Request.")
        req = parse_sheet_request(wb["Request"])
        bang = cell_text(req.get("bang")).lower()
        datasets = column_cfg.get("datasets") or {}
        op_defaults = column_cfg.get("operator_defaults") or {}

        def parse_dataset(dataset_name):
            if dataset_name not in datasets:
                raise RequestError(f"Bảng không hợp lệ: {dataset_name}")
            sheet_name = "Cột Export" if dataset_name == "export" else "Cột Import"
            if sheet_name not in wb.sheetnames:
                raise RequestError(f"Thiếu sheet {sheet_name}.")
            dataset = datasets[dataset_name]
            filters, select_cols, warnings = parse_column_sheet_v6_multi_op(
                wb[sheet_name],
                set(dataset.get("columns") or []),
                op_defaults,
                op_builder,
            )
            return {
                "dataset": dataset,
                "dataset_name": dataset_name,
                "filters": filters,
                "select_cols": select_cols,
                "warnings": warnings,
            }

        if bang == "both":
            export_result = parse_dataset("export")
            import_result = parse_dataset("import")
            if not (
                export_result["filters"]
                or export_result["select_cols"]
                or import_result["filters"]
                or import_result["select_cols"]
            ):
                raise RequestError("Bảng=both nhưng cả 2 sheet đều trống.")
            return {
                "request": req,
                "bang": "both",
                "export": export_result,
                "import": import_result,
            }

        if bang not in datasets:
            raise RequestError(f"Bảng không hợp lệ: {req.get('bang', '')}")
        result = parse_dataset(bang)
        if not result["filters"] and not result["select_cols"]:
            raise RequestError("Chưa chọn cột filter cũng chưa chọn cột lấy về.")
        result["request"] = req
        result["bang"] = bang
        return result
    finally:
        wb.close()


def parse_months(raw):
    text = cell_text(raw).lower()
    if text == "all":
        return [f"{i:02d}" for i in range(1, 13)]
    if not text:
        raise RequestError("Thiếu Tháng.")
    months = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start, end = [p.strip() for p in part.split("-", 1)]
            if not start.isdigit() or not end.isdigit():
                raise RequestError(f"Tháng không hợp lệ: {part}")
            a, b = int(start), int(end)
            if a > b:
                raise RequestError(f"Khoảng tháng không hợp lệ: {part}")
            months.extend(range(a, b + 1))
        else:
            if not part.isdigit():
                raise RequestError(f"Tháng không hợp lệ: {part}")
            months.append(int(part))
    if not months:
        raise RequestError("Thiếu Tháng.")
    bad = [m for m in months if m < 1 or m > 12]
    if bad:
        raise RequestError(f"Tháng ngoài khoảng 1-12: {bad[0]}")
    return [f"{m:02d}" for m in months]


def parse_years(raw):
    text = cell_text(raw).lower()
    if text == "all":
        raise RequestError("Không hỗ trợ all cho Năm. Vui lòng liệt kê năm cụ thể.")
    if not text:
        raise RequestError("Thiếu Năm.")
    years = []
    for part in text.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            start, end = [item.strip() for item in part.split("-", 1)]
            if not start.isdigit() or not end.isdigit():
                raise RequestError(f"Năm không hợp lệ: {part}")
            a, b = int(start), int(end)
            if a > b:
                raise RequestError(f"Khoảng năm không hợp lệ: {part}")
            years.extend(range(a, b + 1))
        else:
            if not part.isdigit():
                raise RequestError(f"Năm không hợp lệ: {part}")
            years.append(int(part))
    if not years:
        raise RequestError("Thiếu Năm.")
    return years


def table_exists(cur, schema, table):
    cur.execute(
        """
        SELECT 1
        FROM information_schema.tables
        WHERE table_schema = %s
          AND table_name = %s
        LIMIT 1
        """,
        (schema, table),
    )
    return cur.fetchone() is not None


def glob_tables(cur, schema, pattern):
    like = pattern.replace("*", "%")
    cur.execute(
        """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = %s
          AND table_name LIKE %s
        ORDER BY table_name
        """,
        (schema, like),
    )
    return [r[0] for r in cur.fetchall()]


def candidate_tables_from_pattern(pattern, today=None):
    today = today or dt.date.today()
    has_year = "{year}" in pattern
    has_month = "{month}" in pattern
    if not has_year and not has_month:
        return [pattern]
    years = range(today.year, today.year - 5, -1) if has_year else [""]
    months = range(12, 0, -1) if has_month else [""]
    candidates = []
    for year in years:
        for month in months:
            values = {
                "year": year,
                "month": f"{month:02d}" if isinstance(month, int) else month,
            }
            candidates.append(pattern.format(**values))
    return candidates


def find_sample_table(cur, schema, pattern, today=None):
    pattern = cell_text(pattern)
    if "*" in pattern and "{year}" not in pattern and "{month}" not in pattern:
        like = pattern.replace("*", "%")
        cur.execute(
            """
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema = %s
              AND table_name LIKE %s
            ORDER BY table_name DESC
            LIMIT 1
            """,
            (schema, like),
        )
        row = cur.fetchone()
        return row[0] if row else None
    for table in candidate_tables_from_pattern(pattern, today=today):
        if table_exists(cur, schema, table):
            return table
    return None


def scan_table_columns(cur, schema, table):
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s
          AND table_name = %s
        ORDER BY ordinal_position
        """,
        (schema, table),
    )
    return [row[0] for row in cur.fetchall()]


def diff_columns(old, new):
    old_set = set(old or [])
    new_set = set(new or [])
    added = [col for col in new if col not in old_set]
    removed = [col for col in old if col not in new_set]
    return added, removed


def confirm_column_overwrite(dataset_name, old_columns, new_columns, input_func=input):
    added, removed = diff_columns(old_columns, new_columns)
    print(
        f"[SCAN] dataset={dataset_name} columns changed: "
        f"+{len(added)} -{len(removed)}"
    )
    if added:
        print(f"[SCAN] added: {', '.join(added)}")
    if removed:
        print(f"[SCAN] removed: {', '.join(removed)}")
    answer = input_func("Overwrite columns in column.yaml? [Y/n] ").strip().lower()
    return answer in ("", "y", "yes")


def scan_columns(column_path=None, cfg=None, dataset_name=None, yes=False, input_func=input, today=None):
    path = Path(column_path or COLUMN_FILE)
    if not path.exists():
        create_column_skeleton(path)
        raise RunnerConfigError(COLUMN_SETUP_MESSAGE)

    column_cfg = load_yaml_file(path, {})
    datasets = column_cfg.get("datasets") or {}
    if not datasets:
        raise RunnerConfigError(COLUMN_SETUP_MESSAGE)
    if dataset_name:
        if dataset_name not in datasets:
            raise RunnerConfigError(f"Dataset không tồn tại trong column.yaml: {dataset_name}")
        scan_names = [dataset_name]
    else:
        scan_names = list(datasets)

    for name in scan_names:
        dataset = datasets.get(name) or {}
        missing = [key for key in ("database", "schema", "tables") if not dataset.get(key)]
        if missing:
            raise RunnerConfigError(f"Dataset {name} thiếu field: {', '.join(missing)}")

    cfg = load_connection_config() if cfg is None else cfg
    messages = []
    updated = False
    for name in scan_names:
        dataset = datasets[name]
        conn = portal.connect(cfg, dataset["database"])
        try:
            if hasattr(conn, "set_client_encoding"):
                conn.set_client_encoding("UTF8")
            cur = conn.cursor()
            try:
                sample_table = find_sample_table(cur, dataset["schema"], dataset["tables"], today=today)
                if not sample_table:
                    message = f"[SCAN] dataset={name} không tìm thấy bảng mẫu cho pattern={dataset['tables']}"
                    print(message)
                    messages.append(message)
                    continue
                columns = scan_table_columns(cur, dataset["schema"], sample_table)
                old_columns = list(dataset.get("columns") or [])
                if old_columns and old_columns != columns and not yes:
                    if not confirm_column_overwrite(name, old_columns, columns, input_func=input_func):
                        message = f"[SCAN] dataset={name} bỏ qua cập nhật columns"
                        print(message)
                        messages.append(message)
                        continue
                dataset["columns"] = columns
                updated = True
                message = (
                    f"[SCAN] dataset={name} sample_table={dataset['schema']}.{sample_table} "
                    f"-> {len(columns)} columns"
                )
                print(message)
                messages.append(message)
            finally:
                cur.close()
        finally:
            conn.close()

    if updated:
        column_cfg["datasets"] = datasets
        column_cfg["operator_defaults"] = column_cfg.get("operator_defaults") or {}
        write_yaml_file(path, column_cfg)
        message = f"[SCAN] operator_defaults giữ nguyên ({len(column_cfg['operator_defaults'])} entries)"
        print(message)
        messages.append(message)
        message = "[SCAN] column.yaml đã cập nhật."
        print(message)
        messages.append(message)
    return messages


def pg_stats_distinct(cur, schema, table, column):
    cur.execute(
        """
        SELECT n_distinct
        FROM pg_stats
        WHERE schemaname = %s
          AND tablename = %s
          AND attname = %s
        """,
        (schema, table, column),
    )
    row = cur.fetchone()
    if not row or row[0] is None:
        return None
    value = float(row[0])
    if value >= 0:
        return int(round(value))

    cur.execute(
        """
        SELECT c.reltuples
        FROM pg_class c
        JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND c.relname = %s
        """,
        (schema, table),
    )
    rel_row = cur.fetchone()
    if not rel_row or rel_row[0] is None:
        return None
    return int(round(abs(value) * float(rel_row[0])))


def sample_distinct_count(cur, schema, table, column, sample_size):
    query = sql.SQL(
        """
        SELECT COUNT(DISTINCT value)
        FROM (
            SELECT {column}::text AS value
            FROM {schema}.{table}
            WHERE {column} IS NOT NULL
            LIMIT %s
        ) sample
        """
    ).format(
        column=sql.Identifier(column),
        schema=sql.Identifier(schema),
        table=sql.Identifier(table),
    )
    cur.execute(query, (sample_size,))
    row = cur.fetchone()
    return int(row[0] or 0)


def sample_text_too_long(cur, schema, table, column, skip_text_length, sample_size):
    query = sql.SQL(
        """
        SELECT AVG(char_length(value))
        FROM (
            SELECT {column}::text AS value
            FROM {schema}.{table}
            WHERE {column} IS NOT NULL
            LIMIT %s
        ) sample
        """
    ).format(
        column=sql.Identifier(column),
        schema=sql.Identifier(schema),
        table=sql.Identifier(table),
    )
    cur.execute(query, (min(sample_size, 50),))
    row = cur.fetchone()
    if not row or row[0] is None:
        return False
    return float(row[0]) > skip_text_length


def distinct_values_for_column(cur, schema, table, column, limit):
    query = sql.SQL(
        """
        SELECT DISTINCT {column}::text AS value
        FROM {schema}.{table}
        WHERE {column} IS NOT NULL
        ORDER BY 1
        LIMIT %s
        """
    ).format(
        column=sql.Identifier(column),
        schema=sql.Identifier(schema),
        table=sql.Identifier(table),
    )
    cur.execute(query, (limit,))
    return [row[0] for row in cur.fetchall()]


def confirm_value_cache_overwrite(dataset_name, old_cache, new_cache, input_func=input):
    old_keys = sorted((old_cache or {}).keys())
    new_keys = sorted((new_cache or {}).keys())
    print(
        f"[SCAN-VALUES] dataset={dataset_name} value_cache changed: "
        f"{len(old_keys)} -> {len(new_keys)} columns"
    )
    answer = input_func("Overwrite cardinality/value cache in column.yaml? [Y/n] ").strip().lower()
    return answer in ("", "y", "yes")


def scan_dataset_values(cur, dataset_name, dataset, settings, column_name=None, today=None):
    sample_table = find_sample_table(cur, dataset["schema"], dataset["tables"], today=today)
    if not sample_table:
        return False, [f"[SCAN-VALUES] dataset={dataset_name} không tìm thấy bảng mẫu cho pattern={dataset['tables']}"]

    columns = list(dataset.get("columns") or [])
    if column_name:
        if column_name not in columns:
            raise RunnerConfigError(f"Cột không tồn tại trong dataset {dataset_name}: {column_name}")
        columns = [column_name]

    skip_columns = set(settings["skip_columns"])
    threshold = settings["threshold"]
    sample_size = settings["sample_size"]
    skip_text_length = settings["skip_text_length"]
    cardinality_cache = dict(dataset.get("cardinality_cache") or {})
    value_cache = dict(dataset.get("value_cache") or {})
    changed = False
    messages = []

    for column in columns:
        if column in skip_columns:
            messages.append(f"[SCAN-VALUES] dataset={dataset_name} column={column} skip configured")
            continue
        if sample_text_too_long(cur, dataset["schema"], sample_table, column, skip_text_length, sample_size):
            cardinality_cache.pop(column, None)
            value_cache.pop(column, None)
            changed = True
            messages.append(f"[SCAN-VALUES] dataset={dataset_name} column={column} skip text long")
            continue

        count = pg_stats_distinct(cur, dataset["schema"], sample_table, column)
        source = "pg_stats"
        if count is None:
            count = sample_distinct_count(cur, dataset["schema"], sample_table, column, sample_size)
            source = "sample"

        cardinality_cache[column] = int(count)
        if count <= threshold:
            values = distinct_values_for_column(cur, dataset["schema"], sample_table, column, threshold + 5)
            value_cache[column] = values[:threshold]
            messages.append(
                f"[SCAN-VALUES] dataset={dataset_name} column={column} distinct={count} source={source} values={len(value_cache[column])}"
            )
        else:
            value_cache.pop(column, None)
            messages.append(
                f"[SCAN-VALUES] dataset={dataset_name} column={column} distinct={count} source={source} values=skip"
            )
        changed = True

    dataset["cardinality_cache"] = cardinality_cache
    dataset["value_cache"] = value_cache
    return changed, messages


def scan_values(column_path=None, cfg=None, dataset_name=None, column_name=None, yes=False, input_func=input, today=None):
    path = Path(column_path or COLUMN_FILE)
    if not path.exists():
        create_column_skeleton(path)
        raise RunnerConfigError(COLUMN_SETUP_MESSAGE)

    raw_column_cfg = load_yaml_file(path, {})
    original_column_cfg = copy.deepcopy(raw_column_cfg)
    column_cfg = ensure_cardinality_schema(raw_column_cfg)
    schema_changed = column_cfg != original_column_cfg
    datasets = column_cfg.get("datasets") or {}
    if not datasets:
        raise RunnerConfigError(COLUMN_SETUP_MESSAGE)
    settings = cardinality_settings(column_cfg)

    if dataset_name:
        if dataset_name not in datasets:
            raise RunnerConfigError(f"Dataset không tồn tại trong column.yaml: {dataset_name}")
        scan_names = [dataset_name]
    else:
        scan_names = list(datasets)

    for name in scan_names:
        dataset = datasets.get(name) or {}
        missing = [key for key in ("database", "schema", "tables", "columns") if not dataset.get(key)]
        if missing:
            raise RunnerConfigError(f"Dataset {name} thiếu field: {', '.join(missing)}")

    cfg = load_connection_config() if cfg is None else cfg
    messages = []
    updated = schema_changed
    proposed = {}
    for name in scan_names:
        original = datasets[name]
        dataset = dict(original)
        conn = portal.connect(cfg, dataset["database"])
        try:
            if hasattr(conn, "set_client_encoding"):
                conn.set_client_encoding("UTF8")
            cur = conn.cursor()
            try:
                changed, dataset_messages = scan_dataset_values(
                    cur,
                    name,
                    dataset,
                    settings,
                    column_name=column_name,
                    today=today,
                )
                for message in dataset_messages:
                    print(message)
                    messages.append(message)
                if changed:
                    proposed[name] = dataset
            finally:
                cur.close()
        finally:
            conn.close()

    for name, dataset in proposed.items():
        old_dataset = datasets[name]
        old_value_cache = old_dataset.get("value_cache") or {}
        old_cardinality_cache = old_dataset.get("cardinality_cache") or {}
        cache_changed = (
            old_value_cache != dataset.get("value_cache", {})
            or old_cardinality_cache != dataset.get("cardinality_cache", {})
        )
        if cache_changed and (yes or confirm_value_cache_overwrite(name, old_value_cache, dataset.get("value_cache"), input_func)):
            datasets[name] = dataset
            updated = True
        elif cache_changed:
            message = f"[SCAN-VALUES] dataset={name} bỏ qua cập nhật cache"
            print(message)
            messages.append(message)

    if updated:
        column_cfg["datasets"] = datasets
        write_yaml_file(path, column_cfg)
        message = "[SCAN-VALUES] column.yaml đã cập nhật."
        print(message)
        messages.append(message)
    return messages


def expand_tables(cur, schema, pattern, request):
    pattern = cell_text(pattern)
    if "{year}" in pattern and not cell_text(request.get("year")):
        raise RequestError("Template cần Năm nhưng request chưa điền.")
    years = [str(year) for year in parse_years(request.get("year"))] if "{year}" in pattern else [""]
    if "{month}" in pattern:
        months = parse_months(request.get("month"))
    else:
        months = [""]

    wanted = []
    if "{year}" in pattern or "{month}" in pattern:
        for year in years:
            for month in months:
                wanted.append(pattern.format(year=year, month=month))
    elif "*" in pattern:
        wanted = glob_tables(cur, schema, pattern)
        if not wanted:
            return [], [pattern]
    else:
        wanted = [pattern]

    existing, missing = [], []
    for table in wanted:
        if table_exists(cur, schema, table):
            existing.append(table)
        else:
            missing.append(table)
    return existing, missing


def column_names(cur, schema, table):
    return [name for name, _dtype in portal.get_columns(cur, schema, table)]


def common_columns(cur, schema, tables):
    per_table = [column_names(cur, schema, table) for table in tables]
    if not per_table:
        return []
    common = set(per_table[0])
    for cols in per_table[1:]:
        common &= set(cols)
    return [col for col in per_table[0] if col in common]


def split_values(raw):
    return [part.strip() for part in cell_text(raw).split(",") if part.strip()]


def parse_split_config(raw, template_split=None):
    text = cell_text(raw)
    if not text and template_split:
        return template_split.get("column"), template_split.get("chars")
    if not text:
        return None, None
    if ":" in text:
        col, chars = [p.strip() for p in text.split(":", 1)]
        if not chars.isdigit() or int(chars) < 1:
            raise RequestError("Tách file theo dạng <cột>:<N>, N phải là số dương.")
        return col, int(chars)
    return text, None


def is_v5_request_file(path):
    wb = load_workbook(path, read_only=True)
    try:
        names = set(wb.sheetnames)
        return "Cột Export" in names or "Cột Import" in names
    finally:
        wb.close()


def normalize_v5_filter_value(op, val, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    return op_builder.portal_value(op, val)


def note_rows_v5(parsed, missing_tables):
    req = parsed["request"]
    rows = [
        ("Bảng", parsed["dataset_name"]),
        ("Người yêu cầu", req.get("user", "")),
        ("Ghi chú / tên request", req.get("request_name", "")),
        ("Năm", req.get("year", "")),
        ("Tháng", req.get("month", "")),
        ("Số filter", len(parsed["filters"])),
        ("Số cột lấy về", len(parsed["select_cols"])),
    ]
    if missing_tables:
        rows.append(("Bảng thiếu", ", ".join(missing_tables)))
    for warning in parsed.get("warnings") or []:
        rows.append(("WARNING", warning))
    return rows


def note_rows_v6(request, dataset_name, dataset_result, missing_tables):
    rows = [
        ("Bảng", dataset_name),
        ("Người yêu cầu", request.get("user", "")),
        ("Ghi chú / tên request", request.get("request_name", "")),
        ("Năm", request.get("year", "")),
        ("Tháng", request.get("month", "")),
        ("Số filter", len(dataset_result["filters"])),
        ("Số cột lấy về", len(dataset_result["select_cols"])),
    ]
    if missing_tables:
        rows.append(("Bảng thiếu", ", ".join(missing_tables)))
    for warning in dataset_result.get("warnings") or []:
        rows.append(("WARNING", warning))
    return rows


def build_where_clause(filters, op_builder=None):
    op_builder = op_builder or load_operator_builder()
    if not filters:
        return sql.SQL(""), []
    fragments = []
    all_params = []
    for item in filters:
        fragment, params = op_builder.build_where(
            item["col"],
            item["op"],
            item["val"],
            item.get("digits"),
        )
        fragments.append(fragment)
        all_params.extend(params)
    return sql.SQL(" WHERE ") + sql.SQL(" AND ").join(fragments), all_params


def build_v6_query(schema, tables, columns, filters, op_builder=None):
    where_sql, params = build_where_clause(filters, op_builder)
    if len(tables) == 1:
        return portal.build_query(schema, tables[0], columns, where_sql), params
    queries = [portal.build_query(schema, table, columns, where_sql, source_label=table) for table in tables]
    return sql.SQL(" UNION ALL ").join(queries), params * len(tables)


def build_jobs_from_v5_request(parsed, cur):
    op_builder = load_operator_builder()
    req = parsed["request"]
    dataset = parsed["dataset"]
    schema = dataset["schema"]
    tables, missing = expand_tables(cur, schema, dataset["tables"], req)
    if not tables:
        raise RequestError(f"Không có bảng nào tồn tại cho pattern: {dataset['tables']}")

    valid_columns = common_columns(cur, schema, tables)
    selected = list(parsed["select_cols"])
    filter_columns = [item["col"] for item in parsed["filters"]]
    needed = selected + filter_columns
    split_col, split_len = parse_split_config(req.get("split"), None)
    if split_col:
        needed.append(split_col)
    bad = [col for col in needed if col not in valid_columns]
    if bad:
        valid = ", ".join(valid_columns)
        raise RequestError(f"Cột không có trong bảng đã chọn: {', '.join(sorted(set(bad)))}. Cột hợp lệ: {valid}")

    filters = [
        (item["col"], item["op"], normalize_v5_filter_value(item["op"], item["val"], op_builder))
        for item in parsed["filters"]
    ]
    state = {
        "db": dataset["database"],
        "schema": schema,
        "tables": tables,
        "cols": selected,
        "filters": filters,
        "split": split_col,
        "split_len": split_len,
        "sort": None,
        "merged": len(tables) > 1,
    }
    jobs = portal.make_jobs(state, cur)
    return dataset["database"], (jobs, note_rows_v5(parsed, missing), selected)


def build_jobs_from_v6_dataset(request, dataset_name, dataset_result, cur, op_builder=None):
    dataset = dataset_result["dataset"]
    if not dataset_result["filters"] and not dataset_result["select_cols"]:
        return dataset["database"], ([], note_rows_v6(request, dataset_name, dataset_result, []), [])

    schema = dataset["schema"]
    tables, missing = expand_tables(cur, schema, dataset["tables"], request)
    if not tables:
        raise RequestError(f"Không có bảng nào tồn tại cho pattern: {dataset['tables']}")

    valid_columns = common_columns(cur, schema, tables)
    selected = list(dataset_result["select_cols"])
    filter_columns = [item["col"] for item in dataset_result["filters"]]
    needed = selected + filter_columns
    bad = [col for col in needed if col not in valid_columns]
    if bad:
        valid = ", ".join(valid_columns)
        raise RequestError(f"Cột không có trong bảng đã chọn: {', '.join(sorted(set(bad)))}. Cột hợp lệ: {valid}")

    query, params = build_v6_query(schema, tables, selected, dataset_result["filters"], op_builder)
    suffix = "merged" if len(tables) > 1 else tables[0]
    name = portal.safe_name(f"{dataset['database']}_{schema}_{suffix}")
    return dataset["database"], ([(name, dataset["database"], query, params)], note_rows_v6(request, dataset_name, dataset_result, missing), selected)


def estimate_mb(rows, columns):
    return rows * max(len(columns), 1) * 15 / (1024 * 1024)


def output_dir(settings):
    folders = settings.get("folders") or {}
    path = Path(folders.get("output") or settings.get("output_dir") or LEGACY_DEFAULT_SETTINGS["output_dir"])
    path.mkdir(parents=True, exist_ok=True)
    return path


def input_dir(settings):
    path = Path(settings.get("input_dir") or LEGACY_DEFAULT_SETTINGS["input_dir"])
    path.mkdir(parents=True, exist_ok=True)
    return path


def approved_dir(settings):
    folders = settings.get("folders") or {}
    if folders.get("approved"):
        path = Path(folders["approved"])
    else:
        path = input_dir(settings)
    path.mkdir(parents=True, exist_ok=True)
    return path


def unique_path(path):
    path = Path(path)
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    for i in range(2, 10000):
        candidate = path.with_name(f"{stem}_{i}{suffix}")
        if not candidate.exists():
            return candidate
    raise RequestError(f"Không tìm được tên file trống cho: {path.name}")


def output_path_for_job(settings, request, suffix, ext):
    ts = dt.datetime.now().strftime("%Y%m%d_%H%M")
    pattern = settings.get("filename_pattern") or DEFAULT_SETTINGS["filename_pattern"]
    base = pattern.format(
        ts=ts,
        user=portal.safe_name(request.get("user", "")),
        request=portal.safe_name(request.get("request_name", "")),
    )
    if suffix:
        base = f"{base}_{portal.safe_name(suffix)}"
    return unique_path(output_dir(settings) / f"{base}{ext}")


def public_job_suffix(job_name, dbname, schema):
    prefix = portal.safe_name(f"{dbname}_{schema}_")
    public = job_name[len(prefix):] if job_name.startswith(prefix) else job_name
    if public.startswith("gop") and "bang_" in public:
        public = public.rsplit("_", 1)[-1]
    return public


def fetch_headers(conn, query, params):
    with conn.cursor() as cur:
        cur.execute(sql.SQL("SELECT * FROM (") + query + sql.SQL(") _h LIMIT 0"), params or None)
        return [d[0] for d in cur.description]


def export_xlsx_v5_with_note(conn, query, params, headers, filepath, notes):
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Data")
    ws.append(headers)
    count = 0
    with conn.cursor(name="bulkex_runner_stream") as cur:
        cur.itersize = 5000
        cur.execute(query, params or None)
        for row in cur:
            ws.append(list(row))
            count += 1
    note_ws = wb.create_sheet("NOTE")
    for row in notes:
        note_ws.append(list(row))
    wb.save(filepath)
    return count


def append_query_to_sheet(conn, ws, query, params, headers):
    if headers:
        ws.append(headers)
    count = 0
    with conn.cursor(name="bulkex_runner_stream") as cur:
        cur.itersize = 5000
        cur.execute(query, params or None)
        for row in cur:
            ws.append(list(row))
            count += 1
    return count


def export_xlsx_v6_both(results, filepath, notes):
    wb = Workbook(write_only=True)
    counts = {}
    for sheet_name, result in results:
        ws = wb.create_sheet(sheet_name)
        counts[sheet_name] = 0
        conn = result["conn"]
        for job in result["jobs"]:
            _name, _dbname, query, params = job
            headers = fetch_headers(conn, query, params)
            counts[sheet_name] += append_query_to_sheet(conn, ws, query, params, headers)
            conn.rollback()
    note_ws = wb.create_sheet("NOTE")
    for row in notes:
        note_ws.append(list(row))
    wb.save(filepath)
    return counts


def export_csv_with_note(conn, query, params, headers, filepath, notes):
    count = 0
    with conn.cursor(name="bulkex_runner_stream") as cur:
        cur.itersize = 5000
        cur.execute(query, params or None)
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in cur:
                writer.writerow(row)
                count += 1
    write_note_txt(Path(filepath).with_suffix(".txt"), notes)
    return count


def write_note_txt(path, notes):
    with open(path, "w", encoding="utf-8") as f:
        for key, value in notes:
            f.write(f"{key}: {value}\n")


def move_finished_file(src, dst):
    dst = unique_path(dst)
    shutil.move(str(src), str(dst))
    return dst


def move_request(path, folder_name):
    path = Path(path)
    target_dir = path.parent / folder_name
    target_dir.mkdir(exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return move_finished_file(path, target_dir / f"{stamp}_{path.name}")


def done_path_for(path, now=None):
    path = Path(path)
    direct = path.with_name(f"{DONE_PREFIX}{path.name}")
    if not direct.exists():
        return direct
    stamp = (now or dt.datetime.now()).strftime("%Y%m%d_%H%M%S")
    return unique_path(path.with_name(f"[DONE {stamp}] {path.name}"))


def mark_done(path):
    return move_finished_file(path, done_path_for(path))


def move_to_error(path, message):
    path = Path(path)
    moved = move_finished_file(path, path.with_name(f"{REJECT_PREFIX}{path.name}"))
    txt = moved.with_suffix(".txt")
    with open(txt, "w", encoding="utf-8") as f:
        f.write(f"File: {path.name}\n")
        f.write(f"Timestamp: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("LỖI:\n")
        for line in str(message).strip().splitlines() or ["Lỗi không xác định"]:
            f.write(f"- {line}\n")
        f.write("\nCách xử lý:\n")
        f.write(f"- Mở file {moved.name}, sửa lại các cột lỗi.\n")
        f.write(f"- Đổi tên bỏ tiền tố {REJECT_PREFIX} hoặc save tên mới, thả lại vào folder.\n")
    return moved


def check_row_limits(rows_by_job, columns, request, settings):
    max_auto = int(settings.get("max_rows_auto", DEFAULT_SETTINGS["max_rows_auto"]))
    max_hard = int(settings.get("max_rows_hard", DEFAULT_SETTINGS["max_rows_hard"]))
    confirmed = cell_text(request.get("large_confirm")).upper() == "YES"
    for rows in rows_by_job:
        mb = estimate_mb(rows, columns)
        if rows > max_hard:
            raise RequestError(
                f"Query ra {rows:,} dòng (ước ~{mb:.1f} MB), vượt ngưỡng cứng. "
                "Vui lòng lọc hẹp hơn hoặc tách request."
            )
        if rows > max_auto and not confirmed:
            raise RequestError(
                f"Query ra {rows:,} dòng (ước ~{mb:.1f} MB). Nếu chắc chắn, "
                "điền ô 'Xác nhận lớn' = YES rồi gửi lại."
            )


def prepare_v6_dataset_jobs(dataset_name, dataset_result, request, cfg, conns, settings, op_builder):
    dataset = dataset_result["dataset"]
    conn = portal.get_conn(conns, cfg, dataset["database"])
    cur = conn.cursor()
    try:
        _dbname, (jobs, base_notes, columns) = build_jobs_from_v6_dataset(
            request, dataset_name, dataset_result, cur, op_builder
        )
        rows_by_job = [portal.count_rows(cur, job[2], job[3]) for job in jobs]
        conn.rollback()
    finally:
        cur.close()
    check_row_limits(rows_by_job, columns, request, settings)
    return {
        "conn": conn,
        "jobs": jobs,
        "notes": base_notes,
        "columns": columns,
        "rows_by_job": rows_by_job,
        "dataset": dataset,
    }


def process_request_file_v6(path, cfg, conns, settings, column_cfg):
    op_builder = load_operator_builder()
    parsed = parse_request_v6(path, column_cfg, op_builder)
    request = parsed["request"]

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    if parsed.get("bang") == "both":
        prepared = []
        all_notes = [("Bảng", "both")]
        for dataset_name, sheet_name in (("export", "Export"), ("import", "Import")):
            result = prepare_v6_dataset_jobs(
                dataset_name, parsed[dataset_name], request, cfg, conns, settings, op_builder
            )
            prepared.append((sheet_name, result))
            all_notes.extend(result["notes"])
            all_notes.append(("Số dòng " + sheet_name, sum(result["rows_by_job"])))
        final_path = output_path_for_job(settings, request, "", ".xlsx")
        tmp_path = unique_path(TMP_DIR / final_path.name)
        counts = export_xlsx_v6_both(prepared, tmp_path, all_notes)
        final = move_finished_file(tmp_path, final_path)
        log_event(f"Exported {final.name}: {counts}")
        moved = mark_done(path)
        log_event(f"Processed request {path.name} -> {moved}")
        return moved

    dataset_name = parsed.get("bang") or parsed.get("dataset_name")
    result = prepare_v6_dataset_jobs(dataset_name, parsed, request, cfg, conns, settings, op_builder)
    dataset = result["dataset"]
    conn = result["conn"]
    jobs = result["jobs"]
    base_notes = result["notes"]
    columns = result["columns"]
    rows_by_job = result["rows_by_job"]
    for index, job in enumerate(jobs):
        name, _dbname, query, params = job
        rows = rows_by_job[index]
        force_csv = rows > XLSX_ROW_LIMIT
        ext = ".csv" if force_csv else ".xlsx"
        suffix = public_job_suffix(name, dataset["database"], dataset["schema"]) if len(jobs) > 1 else ""
        final_path = output_path_for_job(settings, request, suffix, ext)
        tmp_path = unique_path(TMP_DIR / final_path.name)
        notes = list(base_notes)
        notes.append(("Số dòng", rows))
        if rows == 0:
            notes.append(("Gợi ý", "0 dòng: kiểm tra giá trị lọc, đặc biệt MST 10/13 số hoặc dùng bắt đầu bằng."))
        if force_csv:
            notes.append(("Định dạng", "Tự chuyển CSV vì kết quả vượt giới hạn 1,000,000 dòng của Excel."))
        headers = fetch_headers(conn, query, params)
        if force_csv:
            exported = export_csv_with_note(conn, query, params, headers, tmp_path, notes)
            txt_src = tmp_path.with_suffix(".txt")
        else:
            exported = export_xlsx_v5_with_note(conn, query, params, headers, tmp_path, notes)
            txt_src = None
        conn.rollback()
        final = move_finished_file(tmp_path, final_path)
        if txt_src and txt_src.exists():
            move_finished_file(txt_src, final.with_suffix(".txt"))
        log_event(f"Exported {final.name}: {exported} rows")

    moved = mark_done(path)
    log_event(f"Processed request {path.name} -> {moved}")
    return moved


def process_request_file(path, cfg, conns, settings):
    if is_v5_request_file(path):
        return process_request_file_v6(path, cfg, conns, settings, load_column_config())

    raise RequestError(
        "File request không đúng mẫu v5. Vui lòng tạo lại từ request_template.xlsx mới "
        "bằng lệnh: python runner.py --make-template."
    )


process_request_file_v5 = process_request_file_v6


def is_file_stable(path, wait_seconds=5):
    first = Path(path).stat().st_size
    if wait_seconds:
        time.sleep(wait_seconds)
    second = Path(path).stat().st_size
    return first == second


def request_files_v6(settings, stable_wait=5):
    root = approved_dir(settings)
    files = []
    for path in sorted(root.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if path.name.startswith(REJECT_PREFIX):
            continue
        if path.name.startswith(DONE_PREFIX) or path.name.startswith(DONE_TIMESTAMP_PREFIX):
            continue
        if not is_file_stable(path, stable_wait):
            log_event(f"Skip unstable file: {path.name}")
            continue
        files.append(path)
    return files


def request_files(settings, stable_wait=5):
    return request_files_v6(settings, stable_wait=stable_wait)


def done_cleanup_candidates(settings, now=None):
    root = approved_dir(settings)
    freeup_cfg = settings.get("onedrive_freeup") or {}
    delay_hours = float(freeup_cfg.get("approved_delay_hours", DEFAULT_SETTINGS["onedrive_freeup"]["approved_delay_hours"]))
    cutoff = (now or time.time()) - (delay_hours * 3600)
    for path in sorted(root.glob("*.xlsx")):
        if not (path.name.startswith(DONE_PREFIX) or path.name.startswith(DONE_TIMESTAMP_PREFIX)):
            continue
        yield path, path.stat().st_mtime <= cutoff


def output_cleanup_candidates(settings, now=None):
    root = output_dir(settings)
    freeup_cfg = settings.get("onedrive_freeup") or {}
    delay_days = float(freeup_cfg.get("output_delay_days", DEFAULT_SETTINGS["onedrive_freeup"]["output_delay_days"]))
    cutoff = (now or time.time()) - (delay_days * 24 * 3600)
    for path in sorted(root.glob("*.xlsx")):
        yield path, path.stat().st_mtime <= cutoff


def free_up_space(path: Path) -> bool:
    path = Path(path)
    try:
        subprocess.run(["attrib", "+U", "-P", str(path)], check=True, capture_output=True, timeout=10)
        log_event(f"[FREEUP] OK {path.name}")
        return True
    except FileNotFoundError:
        log_event(f"[FREEUP] SKIP {path.name}: 'attrib' not found (non-Windows?)")
        return False
    except subprocess.CalledProcessError as e:
        stderr = e.stderr.decode(errors="ignore") if isinstance(e.stderr, bytes) else str(e.stderr or "")
        log_event(f"[FREEUP] FAIL {path.name}: exit {e.returncode} - {stderr}")
        return False
    except subprocess.TimeoutExpired:
        log_event(f"[FREEUP] TIMEOUT {path.name}: attrib > 10s")
        return False


def cleanup_onedrive(settings=None, now=None):
    settings = load_settings() if settings is None else settings
    freeup_cfg = settings.get("onedrive_freeup") or {}
    if freeup_cfg.get("enabled") is False:
        message = "OneDrive free up disabled in settings"
        print(message)
        log_event(message)
        return {"freed": 0, "skipped": 0, "failed": 0}

    freed = 0
    skipped = 0
    failed = 0
    for path, ready in list(done_cleanup_candidates(settings, now=now)) + list(output_cleanup_candidates(settings, now=now)):
        if not ready:
            skipped += 1
            continue
        if free_up_space(path):
            freed += 1
        else:
            failed += 1

    message = f"[FREEUP] done: freed={freed} skipped={skipped} failed={failed}"
    print(message)
    log_event(message)
    return {"freed": freed, "skipped": skipped, "failed": failed}


def run_once(settings=None, cfg=None, stable_wait=5):
    settings = load_settings() if settings is None else settings
    if settings.get("_deprecated_v5_schema"):
        log_event(V5_SETTINGS_WARNING)
    cfg = load_connection_config() if cfg is None else cfg
    conns = {}
    processed = 0
    try:
        for path in request_files(settings, stable_wait=stable_wait):
            try:
                process_request_file(path, cfg, conns, settings)
                processed += 1
            except psycopg2.OperationalError as e:
                log_event(f"DB down, giữ request {path.name}: {e}")
            except RequestError as e:
                moved = move_to_error(path, str(e))
                log_event(f"Request lỗi {path.name} -> {moved}: {e}")
            except Exception as e:
                moved = move_to_error(path, f"Lỗi xử lý request: {e}")
                log_event(f"Unexpected error {path.name} -> {moved}: {e}")
    finally:
        for conn in conns.values():
            try:
                conn.close()
            except Exception:
                pass
    return processed


def column_union(column_cfg):
    seen = []
    for dataset in (column_cfg.get("datasets") or {}).values():
        for column in dataset.get("columns") or []:
            if column not in seen:
                seen.append(column)
    return seen


def quoted_list_formula(values):
    escaped = [str(value).replace('"', '""') for value in values]
    return '"' + ",".join(escaped) + '"'


def style_table_header(ws, row, columns):
    navy_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    white_font = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def apply_border(ws, min_row, max_row, min_col, max_col):
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border


def add_list_validation(ws, cells, values, allow_blank=True):
    dv = DataValidation(type="list", formula1=quoted_list_formula(values), allow_blank=allow_blank)
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


def add_named_range_validation(ws, cell, range_name, error_style=None):
    dv = DataValidation(type="list", formula1=f"={range_name}", allow_blank=True)
    if error_style:
        dv.errorStyle = error_style
    ws.add_data_validation(dv)
    dv.add(cell)
    return dv


def operator_display_values(op_builder):
    return [display for _key, display in op_builder.display_labels()]


def digits_operator_displays(op_builder):
    return [
        display
        for key, display in op_builder.display_labels()
        if (op_builder.operators.get(key) or {}).get("supports_digits")
    ]


def excel_string_literal(value):
    return '"' + str(value).replace('"', '""') + '"'


def make_values_sheet(wb, column_cfg):
    datasets = column_cfg.get("datasets") or {}
    threshold = cardinality_settings(column_cfg)["threshold"]
    values_by_column = {}
    for dataset in datasets.values():
        cardinality_cache = dataset.get("cardinality_cache") or {}
        value_cache = dataset.get("value_cache") or {}
        for column, values in value_cache.items():
            count = cardinality_cache.get(column)
            if count is None:
                count = len(values or [])
            try:
                count_int = int(count)
            except (TypeError, ValueError):
                continue
            if count_int > threshold or not values:
                continue
            target = values_by_column.setdefault(column, [])
            for value in values:
                text = "" if value is None else str(value)
                if text not in target:
                    target.append(text)

    if not values_by_column:
        return {}

    ws = wb.create_sheet("Values")
    ws.sheet_state = "hidden"
    named_ranges = {}
    for index, (column, values) in enumerate(values_by_column.items(), 1):
        excel_col = get_column_letter(index)
        ws.cell(row=1, column=index, value=column)
        ws.cell(row=1, column=index).font = Font(name="Calibri", size=11, bold=True)
        ws.column_dimensions[excel_col].width = max(14, min(35, len(column) + 4))
        for row_index, value in enumerate(values, 2):
            ws.cell(row=row_index, column=index, value=value)
            ws.cell(row=row_index, column=index).number_format = "@"
        last_row = len(values) + 1
        range_name = f"{column}_values"
        attr_text = f"Values!${excel_col}$2:${excel_col}${last_row}"
        wb.defined_names.add(DefinedName(name=range_name, attr_text=attr_text))
        named_ranges[column] = range_name
    return named_ranges


def setup_request_sheet(ws, column_cfg):
    ws.title = "Request"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.freeze_panes = "B1"
    style_table_header(ws, 1, 1)
    value_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    for row, label in enumerate(REQUEST_V6_LABELS_ORDER, 1):
        header = ws.cell(row=row, column=1, value=label)
        header.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header.font = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
        header.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
        header.border = border
        value = ws.cell(row=row, column=2)
        value.fill = value_fill
        value.border = border
        value.number_format = "@"
        value.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    datasets = list((column_cfg.get("datasets") or {}).keys())
    if datasets:
        dataset_options = list(datasets)
        if "export" in datasets and "import" in datasets and "both" not in dataset_options:
            dataset_options.append("both")
        add_list_validation(ws, "B2", dataset_options, allow_blank=False)
    split_columns = column_union(column_cfg)
    if split_columns:
        add_list_validation(ws, "B5", split_columns, allow_blank=True)
    add_list_validation(ws, "B6", ["YES"], allow_blank=True)
    ws.print_area = "A1:B8"


def add_digits_validation(ws, cells):
    dv = DataValidation(
        type="list",
        formula1='"2, 4, 6, 8, 10, 13"',
        allow_blank=True,
        errorStyle="information",
    )
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


def add_value_dropdowns(ws, columns, named_ranges, value_columns):
    for index, column in enumerate(columns, 2):
        range_name = named_ranges.get(column)
        if not range_name:
            continue
        for value_column in value_columns:
            add_named_range_validation(ws, ws.cell(row=index, column=value_column), range_name, "information")


def setup_column_sheet(ws, title, columns, op_builder, named_ranges):
    ws.title = title
    display_labels = operator_display_values(op_builder)
    ws.append(["Cột"] + display_labels + ["Digits", "Lấy về?"])
    style_table_header(ws, 1, 9)
    ws.freeze_panes = "A2"
    widths = {"A": 35, "H": 10, "I": 12}
    for key in ("B", "C", "D", "E", "F", "G"):
        widths[key] = 14
    for key, width in widths.items():
        ws.column_dimensions[key].width = width

    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    zebra_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    for index, column in enumerate(columns, 2):
        ws.cell(row=index, column=1, value=column)
        if index % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=index, column=col).fill = zebra_fill
        col_cell = ws.cell(row=index, column=1)
        col_cell.fill = gray_fill
        col_cell.font = Font(name="Calibri", size=11, italic=True)
        col_cell.protection = Protection(locked=False)
        for col in range(2, 8):
            ws.cell(row=index, column=col).number_format = "@"
        ws.cell(row=index, column=8).number_format = "@"
        ws.cell(row=index, column=8).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=index, column=9).alignment = Alignment(horizontal="center", vertical="center")

    max_row = max(len(columns) + 1, 2)
    apply_border(ws, 1, max_row, 1, 9)
    add_value_dropdowns(ws, columns, named_ranges, (2, 3))
    add_digits_validation(ws, f"H2:H{max_row}")
    add_list_validation(ws, f"I2:I{max_row}", ["YES", "NO"], allow_blank=True)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws.conditional_formatting.add(
        f"A2:I{max_row}",
        FormulaRule(formula=["COUNTA($B2:$G2)>0"], stopIfTrue=False, fill=yellow_fill),
    )
    inactive_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    inactive_font = Font(color="A6A6A6")
    ws.conditional_formatting.add(
        f"H2:H{max_row}",
        FormulaRule(formula=["AND(ISBLANK($E2), ISBLANK($G2))"], stopIfTrue=False, fill=inactive_fill, font=inactive_font),
    )
    ws.print_area = f"A1:I{max_row}"


def setup_reference_sheet(ws, op_builder):
    ws.title = "Tham chiếu"
    operator_rows = [["Toán tử", "Ý nghĩa", "Cách nhập Giá trị", "Ví dụ", "Có Digits?"]]
    operator_details = {
        "eq": ["Bằng", "Trùng đúng", "1 giá trị hoặc nhiều cách phẩy (tự IN)", "CN, KR", "Không"],
        "in": ["Trong danh sách", "Thuộc list", "Nhiều cách phẩy", "CN, KR, JP", "Không"],
        "between": ["Trong khoảng", "Giữa 2 mốc", "Đúng 2 giá trị cách phẩy", "1000, 5000", "Không"],
        "prefix": ["Bắt đầu bằng", "Prefix", "1 hoặc nhiều cách phẩy", "8306, 8307", "Có"],
        "contains": ["Chứa", "Substring", "1 hoặc nhiều cách phẩy", "laptop, gaming", "Không"],
        "suffix": ["Kết thúc bằng", "Suffix", "1 hoặc nhiều cách phẩy", "AA, BB", "Có"],
    }
    for key, display in op_builder.display_labels():
        row = list(operator_details.get(key) or [display, display, "", "", ""])
        row[0] = display
        row[4] = "Có" if (op_builder.operators.get(key) or {}).get("supports_digits") else row[4]
        operator_rows.append(row)
    sections = operator_rows + [
        [],
        ["Digits", "Ghi chú"],
        [
            "Định nghĩa",
            "Digits = số ký tự bạn muốn match từ bên trái (Bắt đầu bằng) hoặc bên phải (Kết thúc bằng) của cột trong DB.",
        ],
        ["Digits trống", "Không validate độ dài, tool dùng value nguyên."],
        ["Digits có giá trị", "Mọi value bạn điền phải có đúng số ký tự = Digits."],
        [
            "Ví dụ 1",
            "Bắt đầu bằng = 8306, 8307, 8308, 8309; Digits = 4 -> match mọi HS bắt đầu bằng 4 mã này.",
        ],
        [
            "Ví dụ 2",
            "Bắt đầu bằng = 0301234567; Digits = 10 -> match cả MST 10 số và MST 13 số bắt đầu bằng 10 số này.",
        ],
        ["Ví dụ lỗi", "Bắt đầu bằng = 84; Digits = 4 -> LỖI vì value có 2 ký tự, không đủ 4 ký tự."],
        [
            "Gợi ý",
            "2 chapter HS; 4 heading HS; 6 subheading HS; 8 tariff; 10 MST chính hoặc HS national; 13 MST phụ thuộc.",
        ],
        [],
        ["Combine op cùng row", "Ghi chú"],
        ["Nhiều cell op", "Bạn có thể điền value vào nhiều cell op cùng 1 row. Tool tự AND các filter đó lại."],
        ["Ví dụ prefix", "Row ma_so_hang_hoa: Bắt đầu bằng = 84, 85"],
        ["Ví dụ suffix", "Row ma_so_hang_hoa: Kết thúc bằng = 00"],
        ["Ví dụ digits", "Digits = 10 nếu muốn tier national code."],
        ["SQL logic", "(ma_so LIKE '84%' OR '85%') AND ma_so LIKE '%00'"],
        ["Cell op trống", "Op không active."],
        [],
        ["Cột", "Mục đích"],
        ["Cột", "Tên cột DB đã điền sẵn, chỉ là visual reference."],
        ["Bằng/Trong danh sách/...", "Mỗi toán tử là 1 ô riêng. Điền value vào ô muốn dùng."],
        ["Digits", "Độ dài value khi dùng Bắt đầu bằng hoặc Kết thúc bằng."],
        ["Lấy về?", "YES = cột này có trong file kết quả. NO/trống = không."],
        [],
        ["Cú pháp Tháng", "Ý nghĩa"],
        ["03", "Tháng 3"],
        ["01,03,05", "Tháng 1, 3, 5"],
        ["01-06", "Tháng 1 đến 6"],
        ["all", "Cả 12 tháng"],
        [],
        ["Logic quyết định"],
        ["Có Toán tử + Có Giá trị", "Filter WHERE, auto SELECT"],
        ["Có Toán tử + Trống Giá trị", "LỖI: thiếu giá trị"],
        ["Trống Toán tử + Có Giá trị", "Có default thì auto áp default, nếu không thì warning bỏ qua"],
        ["Trống Toán tử + Trống Giá trị", "YES ở Lấy về? thì SELECT, NO/trống thì skip"],
    ]
    for row in sections:
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 56
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 14
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value:
                cell.border = Border(
                    left=Side(style="thin", color="BFBFBF"),
                    right=Side(style="thin", color="BFBFBF"),
                    top=Side(style="thin", color="BFBFBF"),
                    bottom=Side(style="thin", color="BFBFBF"),
                )
    header_rows = [1, 9, 18, 26, 32, 38]
    for row in header_rows:
        style_table_header(ws, row, 5 if row == 1 else 2)


def make_request_template_v5(column_cfg, output_path=None):
    output_path = output_path or REQUEST_TEMPLATE_FILE
    op_builder = load_operator_builder()
    datasets = column_cfg.get("datasets") or {}
    export_columns = list((datasets.get("export") or {}).get("columns") or [])
    import_columns = list((datasets.get("import") or {}).get("columns") or [])
    if not export_columns and not import_columns:
        raise RunnerConfigError("column.yaml chưa có columns. Chạy: python runner.py --scan-columns")

    wb = Workbook()
    named_ranges = make_values_sheet(wb, column_cfg)
    setup_request_sheet(wb.active, column_cfg)
    setup_column_sheet(wb.create_sheet("Cột Export"), "Cột Export", export_columns, op_builder, named_ranges)
    setup_column_sheet(wb.create_sheet("Cột Import"), "Cột Import", import_columns, op_builder, named_ranges)
    if "Values" in wb.sheetnames:
        wb._sheets.append(wb._sheets.pop(wb.sheetnames.index("Values")))
    setup_reference_sheet(wb.create_sheet("Tham chiếu"), op_builder)
    wb.active = 0

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return Path(output_path)


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="SQL BulkEx request runner")
    parser.add_argument("--once", action="store_true", help="Chạy một vòng quét request rồi thoát")
    parser.add_argument("--make-template", action="store_true", help="Sinh request_template.xlsx từ column.yaml")
    parser.add_argument("--scan-columns", action="store_true", help="Quét DB và cập nhật column.yaml")
    parser.add_argument("--scan-values", action="store_true", help="Quét distinct values và cập nhật column.yaml")
    parser.add_argument("--cleanup", action="store_true", help="Free up OneDrive space cho file DONE/output cũ")
    parser.add_argument("--dataset", help="Chỉ scan một dataset trong column.yaml")
    parser.add_argument("--column", help="Chỉ scan một cột khi dùng --scan-values")
    parser.add_argument("--yes", action="store_true", help="Bỏ qua confirm khi cập nhật columns")
    return parser.parse_args(argv)


def main(argv=None):
    configure_stdio()
    args = parse_args(argv)
    if args.scan_columns:
        try:
            scan_columns(dataset_name=args.dataset, yes=args.yes)
            return 0
        except RunnerConfigError as e:
            log_event(str(e))
            print(str(e))
            return 1

    if args.scan_values:
        try:
            scan_values(dataset_name=args.dataset, column_name=args.column, yes=args.yes)
            return 0
        except RunnerConfigError as e:
            log_event(str(e))
            print(str(e))
            return 1

    if args.cleanup:
        try:
            cleanup_onedrive(load_settings())
            return 0
        except RunnerConfigError as e:
            log_event(str(e))
            print(str(e))
            return 1

    if args.make_template:
        try:
            path = make_request_template_v5(load_column_config())
            print(f"Đã sinh file mẫu: {path}")
            return 0
        except RunnerConfigError as e:
            log_event(str(e))
            print(str(e))
            return 1

    settings = load_settings()
    try:
        cfg = load_connection_config()
    except RunnerConfigError as e:
        log_event(str(e))
        print(str(e))
        return 1
    if args.once:
        return 0 if run_once(settings=settings, cfg=cfg) >= 0 else 1

    while True:
        run_once(settings=settings, cfg=cfg)
        time.sleep(int(settings.get("poll_seconds", DEFAULT_SETTINGS["poll_seconds"])))


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nĐã thoát.")
        sys.exit(0)
